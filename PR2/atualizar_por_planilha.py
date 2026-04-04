#!/usr/bin/env python
"""
Entrada por planilha (Excel/PDF) para atualizar e imprimir lotes no CorelDRAW.
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import hmac
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import tempfile
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from difflib import SequenceMatcher
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

import atualizar_ofertas_corel as corel

PRECO_RE = re.compile(r"\d{1,3}(?:[\.,]\d{2})")
PRECO_COM_R_RE = re.compile(r"R\$\s*([\d\s]{1,9})[.,]\s*(\d{2})", re.IGNORECASE)
PRECO_SEM_R_RE = re.compile(r"\b(\d{1,4})\s*[.,]\s*(\d{2})\b")
WHATSAPP_CABECALHO_RE = re.compile(r"^\[\d{1,2}:\d{2},\s*\d{1,2}/\d{1,2}/\d{2,4}\]\s*.+?:\s*$")
VALIDADE_LINHA_RE = re.compile(r"^\s*validade\b", re.IGNORECASE)
DB_APRENDIZADO_ARQUIVO = "aprendizado_produtos.json"
DB_VELOCIDADE_ARQUIVO = "aprendizado_velocidade.json"
DB_CORRETOR_ACENTOS_ARQUIVO = "corretor_acentos.json"
WEB_CONFIG_PAGE_FILE = "configuracao_ofertas_web.html"
WEB_LOGIN_PAGE_FILE = "login_acesso_web.html"
WORD_TOKEN_RE = re.compile(r"[A-Za-zÀ-ÿ]+")
UNIT_OPTIONS = ("Unid.", "Kg", "PCT.", "PACK.", "BDJ.")
MEDIDA_RE = re.compile(
    r"(?P<prefixo>\b(?:c|cx|pct|pack)\s*[/x]?\s*)?"
    r"(?P<numero>\d+(?:[.,]\d+)?)\s*"
    r"(?P<unidade>"
    r"kg|quilo(?:s)?|kilo(?:s)?|kilograma(?:s)?|"
    r"g|gr|grama(?:s)?|mg|"
    r"ml|l|lt|litro(?:s)?|"
    r"unid(?:ades)?|unidade(?:s)?|und|un"
    r")\b",
    re.IGNORECASE,
)
DATA_RE = re.compile(r"\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b")
DATA_SOLTA_RE = re.compile(r"\b\d{1,2}\s+\d{1,2}\s+\d{2,4}\b")
CODIGO_BARRAS_13_RE = re.compile(r"\b\d{13}\b")
CODIGO_INTERNO_RE = re.compile(r"^\d{4,8}$")
DEFAULT_OLLAMA_MODEL = "qwen3.5:0.8b"
DEFAULT_OLLAMA_TIMEOUT_SECONDS = 8.0
DEFAULT_OLLAMA_MAX_ITEMS = 12
DEFAULT_GITHUB_LOG_REPO = "PopularAtacarejo/Placas"
DEFAULT_GITHUB_LOG_PATH = "Ofertas de Validade.json"
DEFAULT_GITHUB_USERS_PATH = "usuarios.json"
DEFAULT_GITHUB_APRENDIZADO_PATH = "Ofertas em Validade/aprendizado_produtos.json"
DEFAULT_GITHUB_LOG_BRANCH = "main"
GITHUB_TOKEN_ENV = "GITHUB_PLACAS_TOKEN"
GITHUB_REPO_ENV = "GITHUB_PLACAS_REPO"
GITHUB_PATH_ENV = "GITHUB_PLACAS_PATH"
GITHUB_BRANCH_ENV = "GITHUB_PLACAS_BRANCH"
GITHUB_APRENDIZADO_PATH_ENV = "GITHUB_PLACAS_APRENDIZADO_PATH"
LOGIN_SESSION_FILE = "sessao_login_24h.json"
LOGIN_SESSION_HOURS = 24
PASSWORD_HASH_SCHEME = "pbkdf2_sha256"
PASSWORD_HASH_ITERATIONS = 390000
PASSWORD_HASH_MIN_ITERATIONS = 150000
PERFIL_ACESSO_LABELS = {
    "administrador": "Administrador",
    "desenvolvedor": "Desenvolvedor",
    "gerador de placas": "Gerador de Placas",
}
PERFIS_COM_ACESSO_GERACAO = set(PERFIL_ACESSO_LABELS.keys())
PERFIS_COM_GESTAO_USUARIOS = {"desenvolvedor"}

_OLLAMA_EXECUTABLE_CACHE: Optional[Path] = None
_OLLAMA_LOOKUP_DONE = False
_OLLAMA_WARNING_EMITIDO = False


def carregar_variaveis_dotenv(caminho_dotenv: Optional[Path] = None) -> None:
    caminho = caminho_dotenv or Path(__file__).with_name(".env")
    if not caminho.exists():
        return
    try:
        linhas = caminho.read_text(encoding="utf-8").splitlines()
    except Exception:
        return
    for linha_bruta in linhas:
        linha = str(linha_bruta or "").strip()
        if not linha or linha.startswith("#"):
            continue
        if linha.lower().startswith("export "):
            linha = linha[7:].strip()
        if "=" not in linha:
            continue
        chave, valor = linha.split("=", 1)
        chave = chave.strip()
        valor = valor.strip()
        if not chave:
            continue
        if len(valor) >= 2 and (
            (valor.startswith('"') and valor.endswith('"'))
            or (valor.startswith("'") and valor.endswith("'"))
        ):
            valor = valor[1:-1]
        if os.environ.get(chave):
            continue
        os.environ[chave] = valor


def encerrar_sessao_web_revisao(estado_revisao: Optional[Dict[str, Any]]) -> None:
    if not isinstance(estado_revisao, dict):
        return
    servidor = estado_revisao.pop("web_server", None)
    thread = estado_revisao.pop("web_server_thread", None)
    if servidor is not None:
        try:
            servidor.shutdown()
        except Exception:
            pass
        try:
            servidor.server_close()
        except Exception:
            pass
    if thread is not None:
        try:
            thread.join(timeout=2.0)
        except Exception:
            pass


def _github_request_json(
    method: str,
    url: str,
    token: Optional[str] = None,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    corpo: Optional[bytes] = None
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "placas-ofertas-automacao",
        "X-GitHub-Api-Version": "2022-11-28",
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if payload is not None:
        corpo = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json; charset=utf-8"

    req = urllib.request.Request(url, data=corpo, headers=headers, method=method.upper())
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            bruto = resp.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        detalhe = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GitHub HTTP {exc.code}: {detalhe or exc.reason}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Falha de rede ao comunicar com o GitHub: {exc.reason}") from exc

    if not bruto.strip():
        return {}
    try:
        parsed = json.loads(bruto)
    except Exception as exc:
        raise RuntimeError("Resposta invalida do GitHub.") from exc
    if not isinstance(parsed, dict):
        raise RuntimeError("Resposta inesperada do GitHub.")
    return parsed


def carregar_json_arquivo_github(
    token: Optional[str],
    repo: str,
    caminho_arquivo: str,
    branch: str,
) -> Tuple[Any, Optional[str]]:
    caminho_url = urllib.parse.quote(caminho_arquivo, safe="/")
    branch_q = urllib.parse.quote(branch, safe="")
    url = f"https://api.github.com/repos/{repo}/contents/{caminho_url}?ref={branch_q}"
    try:
        resposta = _github_request_json("GET", url, token)
    except RuntimeError as exc:
        if "GitHub HTTP 404" in str(exc):
            return [], None
        raise

    sha = str(resposta.get("sha") or "") or None
    conteudo_b64 = str(resposta.get("content") or "").replace("\n", "")
    if not conteudo_b64.strip():
        return [], sha

    try:
        texto = base64.b64decode(conteudo_b64).decode("utf-8-sig")
    except Exception as exc:
        raise RuntimeError("Nao foi possivel decodificar o JSON remoto do GitHub.") from exc

    if not texto.strip():
        return None, sha

    try:
        parsed = json.loads(texto)
    except Exception as exc:
        raise RuntimeError("O arquivo remoto do GitHub nao contem JSON valido.") from exc

    return parsed, sha


def carregar_arquivo_json_github(
    token: Optional[str],
    repo: str,
    caminho_arquivo: str,
    branch: str,
) -> Tuple[List[Dict[str, Any]], Optional[str]]:
    parsed, sha = carregar_json_arquivo_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
    )
    if parsed is None:
        return [], sha
    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)], sha
    raise RuntimeError("O arquivo remoto do GitHub precisa ser um array JSON ou estar vazio.")


def salvar_arquivo_json_github(
    token: str,
    repo: str,
    caminho_arquivo: str,
    branch: str,
    conteudo_json: Any,
    sha_atual: Optional[str],
    mensagem_commit: str,
) -> str:
    caminho_url = urllib.parse.quote(caminho_arquivo, safe="/")
    url = f"https://api.github.com/repos/{repo}/contents/{caminho_url}"
    texto = json.dumps(conteudo_json, ensure_ascii=False, indent=2) + "\n"
    payload: Dict[str, Any] = {
        "message": mensagem_commit,
        "branch": branch,
        "content": base64.b64encode(texto.encode("utf-8")).decode("ascii"),
    }
    if sha_atual:
        payload["sha"] = sha_atual
    resposta = _github_request_json("PUT", url, token, payload=payload)
    content = resposta.get("content")
    if not isinstance(content, dict):
        raise RuntimeError("GitHub nao retornou o SHA atualizado do arquivo.")
    novo_sha = str(content.get("sha") or "").strip()
    if not novo_sha:
        raise RuntimeError("GitHub nao retornou o SHA atualizado do arquivo.")
    return novo_sha


def serializar_lote_para_log(lote: List[Dict[str, str]]) -> List[Dict[str, str]]:
    itens: List[Dict[str, str]] = []
    for indice, item in enumerate(lote, start=1):
        descricao = normalizar_texto(item.get("descricao", ""))
        preco = normalizar_preco_str(str(item.get("preco", "")))
        unidade = normalizar_unidade_saida(str(item.get("unidade", "")))
        validade_oferta = normalizar_data_oferta(item.get("validade_oferta", ""))
        codigo_barras = normalizar_codigo_barras_saida(item.get("codigo_barras", ""))
        usar_codigo_barras = normalizar_flag_codigo_barras_saida(item.get("usar_codigo_barras", False), codigo_barras)
        if not descricao and not preco:
            continue
        registro = {
            "posicao": str(indice),
            "descricao": descricao,
            "preco": preco,
            "unidade": unidade,
        }
        if validade_oferta:
            registro["validade_oferta"] = validade_oferta
        if codigo_barras:
            registro["codigo_barras"] = codigo_barras
            registro["usar_codigo_barras"] = bool(usar_codigo_barras)
        itens.append(registro)
    return itens


def extrair_nomes_produtos_lote(lote: List[Dict[str, str]]) -> List[str]:
    nomes: List[str] = []
    for item in lote:
        descricao = normalizar_texto(item.get("descricao", ""))
        if descricao:
            nomes.append(descricao)
    return nomes


def montar_registro_placa_github(
    numero_placa: int,
    lote: List[Dict[str, str]],
    usuario_logado: Dict[str, Any],
) -> Dict[str, Any]:
    agora = dt.datetime.now().astimezone()
    usuario_id = normalizar_texto(str(usuario_logado.get("usuario") or ""))
    if not usuario_id:
        usuario_id = normalizar_texto(str(usuario_logado.get("nome") or "usuario"))
    return {
        "usuario": usuario_id,
        "data": agora.strftime("%d/%m/%Y"),
        "hora": agora.strftime("%H:%M:%S"),
        "placa": int(numero_placa),
        "produtos": serializar_lote_para_log(lote),
    }


def registrar_placa_concluida_github(
    token: str,
    repo: str,
    caminho_arquivo: str,
    branch: str,
    registro: Dict[str, Any],
) -> None:
    registro_id = str(registro.get("registro_id", "")).strip()
    numero_placa = int(registro.get("placa", 0) or 0)
    data_hora = f"{registro.get('data', '')} {registro.get('hora', '')}".strip()
    usuario_registro = normalizar_texto(str(registro.get("usuario") or ""))
    data_registro = normalizar_texto(str(registro.get("data") or ""))
    hora_registro = normalizar_texto(str(registro.get("hora") or ""))
    produtos_registro = registro.get("produtos")

    for tentativa in range(2):
        registros_atuais, sha_atual = carregar_arquivo_json_github(
            token=token,
            repo=repo,
            caminho_arquivo=caminho_arquivo,
            branch=branch,
        )
        if registro_id and any(str(item.get("registro_id", "")).strip() == registro_id for item in registros_atuais):
            return
        if any(
            normalizar_texto(str(item.get("usuario") or "")) == usuario_registro
            and normalizar_texto(str(item.get("data") or "")) == data_registro
            and normalizar_texto(str(item.get("hora") or "")) == hora_registro
            and int(item.get("placa", 0) or 0) == numero_placa
            and item.get("produtos") == produtos_registro
            for item in registros_atuais
            if isinstance(item, dict)
        ):
            return

        atualizados = list(registros_atuais)
        atualizados.append(registro)
        try:
            salvar_arquivo_json_github(
                token=token,
                repo=repo,
                caminho_arquivo=caminho_arquivo,
                branch=branch,
                conteudo_json=atualizados,
                sha_atual=sha_atual,
                mensagem_commit=f"Registra placa {numero_placa:03d} em {data_hora or 'execucao automatica'}",
            )
            return
        except RuntimeError:
            if tentativa >= 1:
                raise
            time.sleep(0.4)


def normalizar_usuario_login(texto: str) -> str:
    return remover_acentos(normalizar_texto(texto).lower())


def normalizar_email_login(texto: str) -> str:
    return normalizar_texto(texto).strip().lower()


def normalizar_telefone_login(texto: str) -> str:
    return re.sub(r"\D+", "", str(texto or ""))


def _b64_urlsafe_sem_padding(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64_urlsafe_para_bytes(data: str) -> bytes:
    valor = normalizar_texto(data).strip()
    if not valor:
        raise ValueError("base64 vazio")
    padding = "=" * (-len(valor) % 4)
    return base64.urlsafe_b64decode((valor + padding).encode("ascii"))


def gerar_hash_senha(senha_plana: str, iterations: int = PASSWORD_HASH_ITERATIONS) -> str:
    senha = str(senha_plana or "")
    if not senha:
        raise ValueError("Senha vazia nao pode ser criptografada.")
    iter_count = int(iterations)
    if iter_count < PASSWORD_HASH_MIN_ITERATIONS:
        raise ValueError(f"Iteracoes insuficientes ({iter_count}).")
    salt_bytes = os.urandom(16)
    salt = _b64_urlsafe_sem_padding(salt_bytes)
    derivado = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), salt_bytes, iter_count)
    assinatura = _b64_urlsafe_sem_padding(derivado)
    return f"{PASSWORD_HASH_SCHEME}${iter_count}${salt}${assinatura}"


def senha_hash_valida_formato(senha_hash: str) -> bool:
    try:
        esquema, iter_txt, salt_txt, assinatura_txt = str(senha_hash or "").split("$", 3)
        iter_count = int(iter_txt)
        if esquema != PASSWORD_HASH_SCHEME or iter_count < PASSWORD_HASH_MIN_ITERATIONS:
            return False
        _ = _b64_urlsafe_para_bytes(salt_txt)
        _ = _b64_urlsafe_para_bytes(assinatura_txt)
        return True
    except Exception:
        return False


def validar_senha_hash(senha_plana: str, senha_hash: str) -> bool:
    try:
        esquema, iter_txt, salt_txt, assinatura_txt = str(senha_hash or "").split("$", 3)
        if esquema != PASSWORD_HASH_SCHEME:
            return False
        iter_count = int(iter_txt)
        if iter_count < PASSWORD_HASH_MIN_ITERATIONS:
            return False
        salt = _b64_urlsafe_para_bytes(salt_txt)
        assinatura_esperada = _b64_urlsafe_para_bytes(assinatura_txt)
        assinatura_atual = hashlib.pbkdf2_hmac(
            "sha256",
            str(senha_plana or "").encode("utf-8"),
            salt,
            iter_count,
        )
        return hmac.compare_digest(assinatura_atual, assinatura_esperada)
    except Exception:
        return False


def normalizar_perfil_acesso(texto: str) -> str:
    base = remover_acentos(normalizar_texto(texto).lower())
    base = re.sub(r"\s+", " ", base).strip()
    if not base:
        return ""
    aliases = {
        "admin": "administrador",
        "administracao": "administrador",
        "dev": "desenvolvedor",
        "developer": "desenvolvedor",
        "gerador": "gerador de placas",
        "gerador placa": "gerador de placas",
        "gerador de placa": "gerador de placas",
        "gerador-placas": "gerador de placas",
        "gerador_placas": "gerador de placas",
    }
    return aliases.get(base, base)


def rotulo_perfil_acesso(texto: str) -> str:
    perfil_norm = normalizar_perfil_acesso(texto)
    if not perfil_norm:
        return "Gerador de Placas"
    return PERFIL_ACESSO_LABELS.get(perfil_norm, normalizar_texto(texto) or "Gerador de Placas")


def usuario_pode_gerar_placas(usuario: Dict[str, Any]) -> bool:
    perfil_norm = normalizar_perfil_acesso(
        str(usuario.get("perfil") or usuario.get("cargo") or usuario.get("nivel") or "")
    )
    return bool(perfil_norm) and perfil_norm in PERFIS_COM_ACESSO_GERACAO


def usuario_pode_gerenciar_usuarios(usuario: Dict[str, Any]) -> bool:
    perfil_norm = normalizar_perfil_acesso(
        str(usuario.get("perfil") or usuario.get("cargo") or usuario.get("nivel") or "")
    )
    return bool(perfil_norm) and perfil_norm in PERFIS_COM_GESTAO_USUARIOS


def listar_niveis_acesso() -> List[str]:
    return [
        PERFIL_ACESSO_LABELS["gerador de placas"],
        PERFIL_ACESSO_LABELS["desenvolvedor"],
        PERFIL_ACESSO_LABELS["administrador"],
    ]


def serializar_usuario_acesso_repositorio(usuario: Dict[str, Any]) -> Dict[str, Any]:
    usuario_id = normalizar_texto(str(usuario.get("usuario") or ""))
    senha_hash = normalizar_texto(str(usuario.get("senha_hash") or ""))
    nome = normalizar_texto(str(usuario.get("nome") or usuario_id))
    perfil = rotulo_perfil_acesso(str(usuario.get("perfil") or usuario.get("nivel") or "Gerador de Placas"))
    email = normalizar_email_login(str(usuario.get("email") or ""))
    telefone = normalizar_texto(str(usuario.get("telefone") or ""))
    payload: Dict[str, Any] = {
        "usuario": usuario_id,
        "senha_hash": senha_hash,
        "nome": nome or usuario_id,
        "perfil": perfil,
        "ativo": bool(usuario.get("ativo", True)),
    }
    if telefone:
        payload["telefone"] = telefone
    if email:
        payload["email"] = email
    return payload


def criar_usuario_acesso_github(
    token: str,
    repo: str,
    caminho_arquivo: str,
    branch: str,
    criado_por: str,
    novo_usuario: Dict[str, Any],
) -> Dict[str, Any]:
    usuario_id = normalizar_texto(str(novo_usuario.get("usuario") or ""))
    senha_plana = str(novo_usuario.get("senha") or "").strip()
    nome = normalizar_texto(str(novo_usuario.get("nome") or usuario_id))
    nivel_bruto = str(novo_usuario.get("nivel") or novo_usuario.get("perfil") or "")
    nivel_rotulo = rotulo_perfil_acesso(nivel_bruto)
    nivel_norm = normalizar_perfil_acesso(nivel_rotulo)
    email = normalizar_email_login(str(novo_usuario.get("email") or ""))
    telefone = normalizar_texto(str(novo_usuario.get("telefone") or ""))
    ativo = bool(novo_usuario.get("ativo", True))

    if not usuario_id:
        raise ValueError("Informe o usuario para cadastro.")
    if not senha_plana:
        raise ValueError("Informe a senha do novo usuario.")
    if len(senha_plana) < 6:
        raise ValueError("A senha precisa ter pelo menos 6 caracteres.")
    if not nivel_norm or nivel_norm not in PERFIS_COM_ACESSO_GERACAO:
        raise ValueError("Nivel de acesso invalido.")

    bruto, sha = carregar_json_arquivo_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
    )
    usuarios_existentes = normalizar_registros_usuarios_acesso(bruto)
    usuario_norm = normalizar_usuario_login(usuario_id)

    for existente in usuarios_existentes:
        if normalizar_usuario_login(str(existente.get("usuario") or "")) == usuario_norm:
            raise ValueError("Ja existe um usuario com este login.")
        if email and normalizar_email_login(str(existente.get("email") or "")) == email:
            raise ValueError("Ja existe um usuario com este e-mail.")

    novo = {
        "usuario": usuario_id,
        "senha_hash": gerar_hash_senha(senha_plana),
        "nome": nome or usuario_id,
        "perfil": nivel_rotulo,
        "telefone": telefone,
        "email": email,
        "ativo": ativo,
    }
    usuarios_existentes.append(novo)
    usuarios_para_salvar = [serializar_usuario_acesso_repositorio(usuario) for usuario in usuarios_existentes]

    mensagem = f"Cadastro de usuario: {usuario_id} (por {normalizar_texto(criado_por) or 'sistema'})"
    salvar_arquivo_json_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
        conteudo_json=usuarios_para_salvar,
        sha_atual=sha,
        mensagem_commit=mensagem,
    )
    retorno = dict(novo)
    retorno.pop("senha_hash", None)
    return retorno


def usuario_corresponde_login(usuario: Dict[str, Any], identificador: str) -> bool:
    identificador_usuario = normalizar_usuario_login(identificador)
    identificador_email = normalizar_email_login(identificador)
    identificador_telefone = normalizar_telefone_login(identificador)
    if not identificador_usuario and not identificador_email and not identificador_telefone:
        return False

    login_atual = normalizar_usuario_login(str(usuario.get("usuario", "")))
    email_atual = normalizar_email_login(str(usuario.get("email", "")))
    telefone_atual = normalizar_telefone_login(str(usuario.get("telefone", "")))

    if identificador_usuario and login_atual and login_atual == identificador_usuario:
        return True
    if identificador_email and email_atual and email_atual == identificador_email:
        return True
    if identificador_telefone and telefone_atual and telefone_atual == identificador_telefone:
        return True
    return False


def normalizar_registros_usuarios_acesso(bruto: Any) -> List[Dict[str, Any]]:
    if isinstance(bruto, dict):
        candidatos = bruto.get("usuarios")
        if isinstance(candidatos, list):
            bruto = candidatos
        else:
            bruto = [bruto]

    usuarios_saida: List[Dict[str, Any]] = []
    if not isinstance(bruto, list):
        return usuarios_saida

    for item in bruto:
        if not isinstance(item, dict):
            continue
        usuario = normalizar_texto(
            str(item.get("usuario") or item.get("login") or item.get("username") or item.get("user") or "")
        )
        senha_hash = normalizar_texto(str(item.get("senha_hash") or item.get("password_hash") or ""))
        if not usuario or not senha_hash:
            continue
        if not senha_hash_valida_formato(senha_hash):
            continue
        perfil_origem = str(item.get("perfil") or item.get("cargo") or item.get("nivel") or "")
        perfil = rotulo_perfil_acesso(perfil_origem)
        usuarios_saida.append(
            {
                "usuario": usuario,
                "senha_hash": senha_hash,
                "nome": normalizar_texto(str(item.get("nome") or usuario)),
                "perfil": perfil,
                "telefone": normalizar_texto(str(item.get("telefone") or "")),
                "email": normalizar_texto(str(item.get("email") or "")),
                "ativo": bool(item.get("ativo", True)),
            }
        )
    return usuarios_saida


def carregar_usuarios_acesso_github(
    token: Optional[str],
    repo: str,
    caminho_arquivo: str,
    branch: str,
) -> Tuple[List[Dict[str, Any]], str]:
    origem = f"{repo}/{caminho_arquivo}"
    try:
        bruto, _sha = carregar_json_arquivo_github(
            token=token,
            repo=repo,
            caminho_arquivo=caminho_arquivo,
            branch=branch,
        )
        usuarios = normalizar_registros_usuarios_acesso(bruto)
    except Exception as exc:
        print(f"Aviso: nao foi possivel carregar usuarios do GitHub ({exc}).")
        usuarios = []
    return usuarios, origem


def carregar_pagina_login_web() -> str:
    caminho = Path(__file__).with_name(WEB_LOGIN_PAGE_FILE)
    return caminho.read_text(encoding="utf-8")


def caminho_sessao_login() -> Path:
    return Path(__file__).with_name(LOGIN_SESSION_FILE)


def salvar_sessao_login_24h(usuario: Dict[str, Any]) -> None:
    agora = dt.datetime.now().astimezone()
    expira_em = agora + dt.timedelta(hours=LOGIN_SESSION_HOURS)
    payload = {
        "usuario": normalizar_texto(str(usuario.get("usuario") or "")),
        "nome": normalizar_texto(str(usuario.get("nome") or usuario.get("usuario") or "")),
        "perfil": normalizar_texto(str(usuario.get("perfil") or "")),
        "lembrar_ate": expira_em.isoformat(timespec="seconds"),
        "registrado_em": agora.isoformat(timespec="seconds"),
    }
    caminho = caminho_sessao_login()
    caminho.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def limpar_sessao_login_24h() -> None:
    caminho = caminho_sessao_login()
    try:
        if caminho.exists():
            caminho.unlink()
    except Exception:
        pass


def carregar_sessao_login_24h_valida(usuarios: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    caminho = caminho_sessao_login()
    if not caminho.exists():
        return None
    try:
        bruto = json.loads(caminho.read_text(encoding="utf-8"))
    except Exception:
        limpar_sessao_login_24h()
        return None
    if not isinstance(bruto, dict):
        limpar_sessao_login_24h()
        return None

    usuario_salvo = normalizar_usuario_login(str(bruto.get("usuario") or ""))
    lembrar_ate_txt = normalizar_texto(str(bruto.get("lembrar_ate") or ""))
    if not usuario_salvo or not lembrar_ate_txt:
        limpar_sessao_login_24h()
        return None
    try:
        lembrar_ate = dt.datetime.fromisoformat(lembrar_ate_txt)
    except Exception:
        limpar_sessao_login_24h()
        return None
    agora = dt.datetime.now().astimezone()
    if lembrar_ate.tzinfo is None:
        lembrar_ate = lembrar_ate.replace(tzinfo=agora.tzinfo)
    if lembrar_ate <= agora:
        limpar_sessao_login_24h()
        return None

    for usuario in usuarios:
        if not bool(usuario.get("ativo", True)):
            continue
        if normalizar_usuario_login(str(usuario.get("usuario", ""))) == usuario_salvo:
            if not usuario_pode_gerar_placas(usuario):
                limpar_sessao_login_24h()
                return None
            usuario_saida = dict(usuario)
            usuario_saida["sessao_24h"] = True
            usuario_saida["sessao_expira_em"] = lembrar_ate.isoformat(timespec="seconds")
            return usuario_saida

    limpar_sessao_login_24h()
    return None


def autenticar_login(
    usuarios: List[Dict[str, Any]],
    usuario_digitado: str,
    senha_digitada: str,
) -> Optional[Dict[str, Any]]:
    senha_norm = str(senha_digitada).strip()
    if not normalizar_texto(usuario_digitado) or not senha_norm:
        return None
    for usuario in usuarios:
        if not bool(usuario.get("ativo", True)):
            continue
        senha_hash = normalizar_texto(str(usuario.get("senha_hash") or ""))
        if not senha_hash:
            continue
        if usuario_corresponde_login(usuario, usuario_digitado) and validar_senha_hash(senha_norm, senha_hash):
            return dict(usuario)
    return None


def serializar_usuario_logado(usuario_logado: Optional[Dict[str, Any]]) -> Dict[str, str]:
    usuario = usuario_logado if isinstance(usuario_logado, dict) else {}
    nome = normalizar_texto(str(usuario.get("nome") or usuario.get("usuario") or ""))
    perfil = normalizar_texto(str(usuario.get("perfil") or ""))
    usuario_id = normalizar_texto(str(usuario.get("usuario") or ""))
    meta = (
        f"Sessao de 24 horas ate {normalizar_texto(str(usuario.get('sessao_expira_em') or ''))}"
        if normalizar_texto(str(usuario.get("sessao_expira_em") or ""))
        else "Acesso liberado para esta sessao."
    )
    if not nome and not usuario_id:
        try:
            bruto = json.loads(caminho_sessao_login().read_text(encoding="utf-8"))
        except Exception:
            bruto = {}
        if isinstance(bruto, dict):
            usuario_id = usuario_id or normalizar_texto(str(bruto.get("usuario") or ""))
            nome = nome or normalizar_texto(str(bruto.get("nome") or bruto.get("usuario") or ""))
            perfil = perfil or normalizar_texto(str(bruto.get("perfil") or ""))
            lembrar_ate = normalizar_texto(str(bruto.get("lembrar_ate") or ""))
            if lembrar_ate:
                meta = f"Sessao de 24 horas ate {lembrar_ate}"
    if not nome and not usuario_id:
        return {
            "usuario": "",
            "nome": "Sessao encerrada",
            "perfil": "Entre com outro usuario",
            "meta": "Entre com usuario e senha para continuar a producao.",
            "can_manage_users": False,
            "access_levels": listar_niveis_acesso(),
        }
    return {
        "usuario": usuario_id,
        "nome": nome or usuario_id or "Usuario",
        "perfil": perfil or "Sessao ativa",
        "meta": meta,
        "can_manage_users": bool(usuario_pode_gerenciar_usuarios(usuario)),
        "access_levels": listar_niveis_acesso(),
    }


class EntradaInvalida(Exception):
    pass


class ProducaoCancelada(Exception):
    pass


def normalizar_texto(valor: object) -> str:
    if valor is None:
        return ""
    txt = str(valor).replace("\n", " ").replace("\r", " ").strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt


def normalizar_preco_str(valor: str) -> str:
    txt = normalizar_texto(valor)
    if not txt:
        return ""

    m_r = PRECO_COM_R_RE.search(txt)
    if m_r:
        inteiro = re.sub(r"\D", "", m_r.group(1))
        dec = re.sub(r"\D", "", m_r.group(2))
        if inteiro and len(dec) >= 2:
            return f"{inteiro},{dec[:2]}"

    m = PRECO_SEM_R_RE.search(txt.replace(" ", ""))
    if m:
        inteiro = re.sub(r"\D", "", m.group(1))
        dec = re.sub(r"\D", "", m.group(2))
        if inteiro and len(dec) >= 2:
            return f"{inteiro},{dec[:2]}"

    txt_sem_espaco = txt.replace("R$", "").replace(" ", "").replace(".", ",")
    m_old = PRECO_RE.search(txt_sem_espaco)
    if m_old:
        return m_old.group(0).replace(".", ",")
    return ""


def normalizar_data_oferta(valor: object) -> str:
    if valor is None:
        return ""
    if isinstance(valor, dt.datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, dt.date):
        return valor.strftime("%d/%m/%Y")

    txt = normalizar_texto(valor)
    if not txt:
        return ""

    m_iso = re.search(r"\b(\d{4})-(\d{1,2})-(\d{1,2})(?:\b|[ T])", txt)
    if m_iso:
        ano, mes, dia = m_iso.groups()
        return f"{int(dia):02d}/{int(mes):02d}/{int(ano):04d}"

    m_data = re.search(r"\b(\d{1,2})[/-](\d{1,2})(?:[/-](\d{2,4}))?\b", txt)
    if m_data:
        dia, mes, ano = m_data.groups()
        ano_int = int(ano) if ano else dt.date.today().year
        if ano_int < 100:
            ano_int += 2000
        return f"{int(dia):02d}/{int(mes):02d}/{int(ano_int):04d}"

    m_solta = re.search(r"\b(\d{1,2})\s+(\d{1,2})\s+(\d{2,4})\b", txt)
    if m_solta:
        dia, mes, ano = m_solta.groups()
        ano_int = int(ano)
        if ano_int < 100:
            ano_int += 2000
        return f"{int(dia):02d}/{int(mes):02d}/{int(ano_int):04d}"

    return ""


def extrair_digitos_codigo_barras(valor: object) -> str:
    return re.sub(r"\D", "", normalizar_texto(valor))


def normalizar_codigo_barras_saida(valor: object) -> str:
    digitos = extrair_digitos_codigo_barras(valor)
    if len(digitos) == 13:
        return digitos
    return ""


def normalizar_flag_codigo_barras_saida(valor: object, codigo_barras: object = "") -> bool:
    if isinstance(valor, bool):
        return valor
    txt = normalizar_texto(valor).strip().lower()
    if txt in {"0", "false", "falso", "nao", "não", "off", "desativado"}:
        return False
    if txt in {"1", "true", "verdadeiro", "sim", "on", "ativado"}:
        return True
    return bool(normalizar_codigo_barras_saida(codigo_barras))


def parece_data_texto(valor: object) -> bool:
    return bool(normalizar_data_oferta(valor))


def extrair_data_celulas(cells: Sequence[object]) -> str:
    for cell in cells:
        data = normalizar_data_oferta(cell)
        if data:
            return data
    return ""


def extrair_codigo_barras_celulas(cells: Sequence[str]) -> str:
    for cell in cells:
        codigo = normalizar_codigo_barras_saida(cell)
        if codigo:
            return codigo
    return ""


def extrair_descricao_tabela(cells: Sequence[str]) -> str:
    candidatos: List[str] = []
    for cell in cells:
        txt = normalizar_texto(cell)
        if not txt:
            continue
        txt_norm = remover_acentos(txt).lower()
        if txt_norm.startswith(
            (
                "descricao",
                "descrição",
                "descricao do produto",
                "descrição do produto",
                "produto",
                "item",
                "codigo",
                "código",
                "acao",
                "ação",
                "data",
            )
        ):
            continue
        if normalizar_preco_str(txt):
            continue
        if parece_data_texto(txt):
            continue
        if normalizar_codigo_barras_saida(txt):
            continue
        if CODIGO_INTERNO_RE.fullmatch(txt):
            continue
        if not re.search(r"[A-Za-zÀ-ÿ]", txt):
            continue
        candidatos.append(txt)

    if not candidatos:
        return ""
    return max(candidatos, key=lambda item: (len(re.sub(r"\W+", "", item)), len(item)))


def extrair_preco_e_descricao_da_linha(linha: str) -> Tuple[str, str]:
    txt = normalizar_texto(linha)
    if not txt:
        return "", ""

    m_r = PRECO_COM_R_RE.search(txt)
    if m_r:
        inteiro = re.sub(r"\D", "", m_r.group(1))
        dec = re.sub(r"\D", "", m_r.group(2))
        if inteiro and len(dec) >= 2:
            preco = f"{inteiro},{dec[:2]}"
            descricao = txt[: m_r.start()].strip(" -;|\t")
            return descricao, preco

    m = PRECO_SEM_R_RE.search(txt.replace(" ", ""))
    if m:
        inteiro = re.sub(r"\D", "", m.group(1))
        dec = re.sub(r"\D", "", m.group(2))
        if inteiro and len(dec) >= 2:
            preco = f"{inteiro},{dec[:2]}"
            # Fallback para caso raro sem R$, usa texto antes do primeiro trecho com decimal.
            m_old = PRECO_RE.search(txt)
            if m_old:
                descricao = txt[: m_old.start()].strip(" -;|\t")
            else:
                descricao = txt
            return descricao, preco

    return "", ""


def extrair_preco_das_celulas(cells: Sequence[str]) -> str:
    # Caso comum: coluna "R$" separada do valor ("33,98")
    for i, c in enumerate(cells):
        c_norm = normalizar_texto(c).lower().replace(" ", "")
        if c_norm in ("r$", "r"):
            candidatos: List[str] = []
            if i + 1 < len(cells):
                candidatos.append(normalizar_texto(cells[i + 1]))
            if i + 2 < len(cells):
                candidatos.append(
                    f"{normalizar_texto(cells[i + 1])} {normalizar_texto(cells[i + 2])}".strip()
                )
            for cand in candidatos:
                p = normalizar_preco_str(cand)
                if p:
                    return p

    # Caso combinado na mesma celula/linha
    linha = " ".join(normalizar_texto(c) for c in cells if normalizar_texto(c))
    _, p = extrair_preco_e_descricao_da_linha(linha)
    if p:
        return p
    return normalizar_preco_str(linha)


def detectar_colunas(cabecalho: Sequence[str]) -> Optional[Dict[str, int]]:
    desc_idx = None
    unid_idx = None
    preco_idx = None
    data_idx = None

    for i, col in enumerate(cabecalho):
        c = col.lower()
        if desc_idx is None and any(k in c for k in ("descricao", "descri", "produto", "item")):
            desc_idx = i
        if unid_idx is None and any(k in c for k in ("unid", "unidade", "medida")):
            unid_idx = i
        if preco_idx is None and any(k in c for k in ("preco", "valor", "venda", "oferta")):
            preco_idx = i
        if data_idx is None and any(k in c for k in ("data", "validade", "venc", "ate", "até")):
            data_idx = i

    if desc_idx is None or preco_idx is None:
        return None

    return {
        "descricao": desc_idx,
        "unidade": unid_idx if unid_idx is not None else -1,
        "preco": preco_idx,
        "validade_oferta": data_idx if data_idx is not None else -1,
    }


def deduplicar_produtos(produtos: List[Dict[str, str]]) -> List[Dict[str, str]]:
    unicos: List[Dict[str, str]] = []
    vistos: Dict[Tuple[str, str, str], int] = {}
    for p in produtos:
        desc = normalizar_texto(p.get("descricao", ""))
        unid = normalizar_texto(p.get("unidade", "Unid.")) or "Unid."
        preco = normalizar_preco_str(str(p.get("preco", "")))
        validade_oferta = normalizar_data_oferta(p.get("validade_oferta", ""))
        codigo_barras = normalizar_codigo_barras_saida(p.get("codigo_barras", ""))
        usar_codigo_barras = normalizar_flag_codigo_barras_saida(p.get("usar_codigo_barras", False), codigo_barras)
        if not desc or not preco:
            continue

        chave = (chave_produto(desc), preco, unid.lower())
        if chave in vistos:
            idx_existente = vistos[chave]
            if codigo_barras and not normalizar_codigo_barras_saida(unicos[idx_existente].get("codigo_barras", "")):
                unicos[idx_existente]["codigo_barras"] = codigo_barras
            if codigo_barras and usar_codigo_barras:
                unicos[idx_existente]["usar_codigo_barras"] = True
            if validade_oferta and not normalizar_data_oferta(unicos[idx_existente].get("validade_oferta", "")):
                unicos[idx_existente]["validade_oferta"] = validade_oferta
            continue

        vistos[chave] = len(unicos)
        unicos.append(
            {
                "descricao": desc,
                "unidade": unid,
                "preco": preco,
                "validade_oferta": validade_oferta,
                "codigo_barras": codigo_barras,
                "usar_codigo_barras": bool(usar_codigo_barras),
            }
        )
    return unicos


def extrair_excel(caminho: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise EntradaInvalida(
            "Dependencia faltando para Excel: openpyxl. Instale com 'pip install openpyxl'."
        ) from exc

    if caminho.suffix.lower() == ".xls":
        raise EntradaInvalida("Arquivo .xls nao suportado diretamente. Salve como .xlsx e tente novamente.")

    wb = load_workbook(filename=str(caminho), data_only=True)
    ws = wb.active

    linhas: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        vals = [normalizar_texto(v) for v in row]
        if any(vals):
            linhas.append(vals)

    if not linhas:
        raise EntradaInvalida("Planilha vazia.")

    mapa_colunas = None
    inicio_dados = 0

    for i in range(min(10, len(linhas))):
        cand = detectar_colunas(linhas[i])
        if cand:
            mapa_colunas = cand
            inicio_dados = i + 1
            break

    produtos: List[Dict[str, str]] = []

    if mapa_colunas:
        for row in linhas[inicio_dados:]:
            d_idx = mapa_colunas["descricao"]
            p_idx = mapa_colunas["preco"]
            u_idx = mapa_colunas["unidade"]
            v_idx = mapa_colunas.get("validade_oferta", -1)

            if d_idx >= len(row) or p_idx >= len(row):
                continue

            descricao = row[d_idx]
            preco = ""
            candidatos_preco: List[str] = [row[p_idx]]
            if p_idx + 1 < len(row):
                candidatos_preco.append(f"{row[p_idx]} {row[p_idx + 1]}")
                candidatos_preco.append(row[p_idx + 1])
            preco = extrair_preco_das_celulas(candidatos_preco)
            unidade = row[u_idx] if (u_idx >= 0 and u_idx < len(row)) else "Unid."
            validade_oferta = row[v_idx] if (v_idx >= 0 and v_idx < len(row)) else extrair_data_celulas(row)
            codigo_barras = extrair_codigo_barras_celulas(row)

            if not descricao or not preco:
                continue

            produtos.append(
                {
                    "descricao": descricao,
                    "unidade": unidade or "Unid.",
                    "preco": preco,
                    "validade_oferta": normalizar_data_oferta(validade_oferta),
                    "codigo_barras": codigo_barras,
                    "usar_codigo_barras": bool(codigo_barras),
                }
            )
    else:
        for row in linhas:
            cells = [c for c in row if c]
            if len(cells) < 2:
                continue

            preco = extrair_preco_das_celulas(cells)
            preco_idx = -1
            if preco:
                for i, c in enumerate(cells):
                    if normalizar_preco_str(c) == preco:
                        preco_idx = i
                        break

            if not preco:
                continue

            descricao = cells[0]
            unidade = "Unid."
            if preco_idx + 1 < len(cells) and "unid" in cells[preco_idx + 1].lower():
                unidade = cells[preco_idx + 1]
            validade_oferta = extrair_data_celulas(cells)
            codigo_barras = extrair_codigo_barras_celulas(cells)

            produtos.append(
                {
                    "descricao": descricao,
                    "unidade": unidade,
                    "preco": preco,
                    "validade_oferta": validade_oferta,
                    "codigo_barras": codigo_barras,
                    "usar_codigo_barras": bool(codigo_barras),
                }
            )

    produtos = deduplicar_produtos(produtos)
    if not produtos:
        raise EntradaInvalida(
            "Nao foi possivel extrair produtos do Excel. "
            "Use colunas com Descricao, Unidade e Preco."
        )

    return produtos


def parse_linha_pdf(linha: str) -> Optional[Dict[str, str]]:
    linha = normalizar_texto(linha)
    if not linha:
        return None

    antes, preco = extrair_preco_e_descricao_da_linha(linha)
    if not preco:
        return None

    m_old = PRECO_RE.search(linha)
    fim_preco = m_old.end() if m_old else len(linha)
    depois = linha[fim_preco:].strip(" -;|\t")
    codigo_barras = ""
    match_codigo = CODIGO_BARRAS_13_RE.search(antes)
    if match_codigo:
        codigo_barras = match_codigo.group(0)
        antes = (antes[: match_codigo.start()] + " " + antes[match_codigo.end() :]).strip()
    antes = re.sub(r"^\s*(?:\d{4,13}\s+){1,4}", "", antes).strip()

    if not antes:
        return None

    unidade = "Unid."
    if "unid" in linha.lower() and depois and not normalizar_data_oferta(depois):
        unidade = depois
    validade_oferta = normalizar_data_oferta(depois) or normalizar_data_oferta(linha)

    return {
        "descricao": antes,
        "unidade": unidade,
        "preco": preco,
        "validade_oferta": validade_oferta,
        "codigo_barras": codigo_barras,
        "usar_codigo_barras": bool(codigo_barras),
    }


def extrair_pdf(caminho: Path) -> List[Dict[str, str]]:
    try:
        import pdfplumber  # type: ignore
    except ImportError as exc:
        raise EntradaInvalida(
            "Dependencia faltando para PDF: pdfplumber. Instale com 'pip install pdfplumber'."
        ) from exc

    produtos: List[Dict[str, str]] = []

    with pdfplumber.open(str(caminho)) as pdf:
        for page in pdf.pages:
            tabelas = page.extract_tables() or []
            for tabela in tabelas:
                for row in tabela:
                    cells = [normalizar_texto(c) for c in (row or []) if normalizar_texto(c)]
                    if not cells:
                        continue

                    preco = extrair_preco_das_celulas(cells)

                    if not preco:
                        continue

                    descricao = extrair_descricao_tabela(cells)
                    if not descricao or descricao.lower().startswith(("descricao", "produto", "item")):
                        continue

                    unidade = "Unid."
                    for c in cells:
                        if "unid" in c.lower() or c.lower() == "kg":
                            unidade = c
                            break
                    if unidade == "Unid.":
                        unidade = inferir_unidade_por_descricao(descricao)

                    produtos.append(
                        {
                            "descricao": descricao,
                            "unidade": unidade,
                            "preco": preco,
                            "validade_oferta": extrair_data_celulas(cells),
                            "codigo_barras": extrair_codigo_barras_celulas(cells),
                            "usar_codigo_barras": bool(extrair_codigo_barras_celulas(cells)),
                        }
                    )

            texto = page.extract_text() or ""
            for linha in texto.splitlines():
                p = parse_linha_pdf(linha)
                if not p:
                    continue
                if p["descricao"].lower().startswith(("descricao", "produto", "item")):
                    continue
                produtos.append(p)

    produtos = deduplicar_produtos(produtos)
    if not produtos:
        raise EntradaInvalida(
            "Nao foi possivel extrair produtos do PDF. "
            "Prefira PDF com tabela clara (Descricao/Preco/Unidade)."
        )

    return produtos


def ler_texto_arquivo(caminho: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return caminho.read_text(encoding=encoding)
        except Exception:
            continue
    raise EntradaInvalida(f"Nao foi possivel ler o arquivo de texto: {caminho}")


def inferir_unidade_por_descricao(descricao: str) -> str:
    txt = remover_acentos(normalizar_espacamento_produto(descricao).lower())
    if not txt:
        return "Unid."
    if re.search(r"\b(unid(?:ade)?|und|un)\b", txt):
        return "Unid."
    if re.search(r"\b(ovos?|bandeja|bdj|pack|pct|pacote)\b", txt):
        return "Unid."
    if re.search(r"\b\d+(?:[.,]\d+)?\s*(ml|g|gr|kg|l|lt)\b", txt):
        return "Unid."
    if re.search(r"(?<!\d)\bkg\b", txt):
        return "Kg"
    return "Unid."


def extrair_texto_bruto_lista(texto: str) -> List[Dict[str, str]]:
    produtos: List[Dict[str, str]] = []
    descricao_buffer: List[str] = []

    for linha_bruta in texto.splitlines():
        linha = normalizar_texto(linha_bruta)
        if not linha:
            continue
        if WHATSAPP_CABECALHO_RE.match(linha):
            descricao_buffer.clear()
            continue
        if VALIDADE_LINHA_RE.match(linha):
            continue

        descricao_inline, preco_inline = extrair_preco_e_descricao_da_linha(linha)
        preco_linha = preco_inline or normalizar_preco_str(linha)
        if preco_linha:
            descricao = normalizar_texto(descricao_inline or " ".join(descricao_buffer))
            descricao_buffer.clear()
            if not descricao:
                continue
            produtos.append(
                {
                    "descricao": descricao,
                    "unidade": inferir_unidade_por_descricao(descricao),
                    "preco": preco_linha,
                    "codigo_barras": "",
                    "usar_codigo_barras": False,
                }
            )
            continue

        descricao_buffer.append(linha)

    produtos = deduplicar_produtos(produtos)
    if not produtos:
        raise EntradaInvalida(
            "Nao foi possivel extrair produtos do texto. "
            "Use linhas com descricao e preco no formato da mensagem."
        )
    return produtos


def extrair_texto_lista(caminho: Path) -> List[Dict[str, str]]:
    return extrair_texto_bruto_lista(ler_texto_arquivo(caminho))


def extrair_produtos(caminho: Path) -> List[Dict[str, str]]:
    ext = caminho.suffix.lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return extrair_excel(caminho)
    if ext == ".pdf":
        return extrair_pdf(caminho)
    if ext in (".txt", ".md"):
        return extrair_texto_lista(caminho)
    raise EntradaInvalida("Formato nao suportado. Use .xlsx, .xlsm, .pdf, .txt ou .md")


def selecionar_entrada_tela() -> Optional[Dict[str, str]]:
    try:
        import tkinter as tk
        from tkinter import filedialog, messagebox, scrolledtext, ttk
    except Exception:
        return None

    resultado: Dict[str, str] = {}
    root = tk.Tk()
    root.title("Selecionar Entrada")
    root.geometry("1180x780")
    root.minsize(980, 680)
    configurar_estilo_interface(root, ttk)

    cabecalho = ttk.Frame(root, style="Hero.TFrame", padding=(24, 22))
    cabecalho.pack(fill="x")
    hero_grid = ttk.Frame(cabecalho, style="Hero.TFrame")
    hero_grid.pack(fill="x")
    hero_grid.columnconfigure(0, weight=1)
    hero_grid.columnconfigure(1, weight=0)

    hero_texto = ttk.Frame(hero_grid, style="Hero.TFrame")
    hero_texto.grid(row=0, column=0, sticky="nsew")
    ttk.Label(hero_texto, text="Entrada da automacao", style="HeroTitle.TLabel").pack(anchor="w")
    ttk.Label(hero_texto, text="ETAPA 01  •  CAPTURA DE OFERTAS", style="HeroMeta.TLabel").pack(anchor="w", pady=(2, 0))
    ttk.Label(
        hero_texto,
        text=(
            "Escolha um arquivo Excel, PDF, TXT ou cole o texto bruto da mensagem. "
            "O sistema ignora linhas de validade, corrige descricoes e prepara a lista para a producao."
        ),
        style="HeroText.TLabel",
        wraplength=760,
        justify="left",
    ).pack(anchor="w", pady=(10, 0))

    hero_info = ttk.Frame(hero_grid, style="Card.TFrame", padding=(16, 14))
    hero_info.grid(row=0, column=1, sticky="ne", padx=(18, 0))
    ttk.Label(hero_info, text="Fluxo rapido", style="CardTitle.TLabel").pack(anchor="w")
    ttk.Label(hero_info, text="1. importar\n2. revisar\n3. produzir", style="CardValue.TLabel", justify="left").pack(anchor="w", pady=(6, 0))
    ttk.Label(hero_info, text="Entrada flexivel com apoio de IA local e memoria de aprendizado.", style="SectionText.TLabel", wraplength=240, justify="left").pack(anchor="w", pady=(8, 0))

    corpo = ttk.Frame(root, style="Page.TFrame", padding=(20, 18))
    corpo.pack(fill="both", expand=True)
    corpo.columnconfigure(0, weight=0)
    corpo.columnconfigure(1, weight=1)
    corpo.rowconfigure(1, weight=1)

    modo_var = tk.StringVar(value="arquivo")
    arquivo_var = tk.StringVar(value="")
    status_var = tk.StringVar(value="Modo arquivo ativo.")

    lateral = ttk.Frame(corpo, style="Page.TFrame")
    lateral.grid(row=0, column=0, rowspan=2, sticky="ns", padx=(0, 18))

    card_modo = ttk.Frame(lateral, style="Card.TFrame", padding=(16, 16))
    card_modo.pack(fill="x")
    ttk.Label(card_modo, text="Origem da entrada", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        card_modo,
        text="Escolha se voce vai importar um arquivo ou colar o conteudo bruto da oferta.",
        style="SectionText.TLabel",
        wraplength=280,
        justify="left",
    ).pack(anchor="w", pady=(4, 12))
    ttk.Radiobutton(card_modo, text="📂 Arquivo", value="arquivo", variable=modo_var, style="Picker.TRadiobutton").pack(fill="x")
    ttk.Radiobutton(card_modo, text="📝 Texto bruto", value="texto", variable=modo_var, style="Picker.TRadiobutton").pack(fill="x", pady=(8, 0))
    ttk.Label(card_modo, textvariable=status_var, style="Status.TLabel", wraplength=280, justify="left").pack(anchor="w", pady=(14, 0))

    card_dicas = ttk.Frame(lateral, style="Card.TFrame", padding=(16, 16))
    card_dicas.pack(fill="x", pady=(14, 0))
    ttk.Label(card_dicas, text="Formatos aceitos", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        card_dicas,
        text="• Excel (.xlsx, .xlsm)\n• PDF\n• TXT e MD\n• Texto colado diretamente",
        style="SectionText.TLabel",
        justify="left",
    ).pack(anchor="w", pady=(6, 0))
    ttk.Label(
        card_dicas,
        text="Dica: prefira o modo arquivo quando a oferta vier de planilha ou PDF estruturado.",
        style="SectionText.TLabel",
        wraplength=280,
        justify="left",
    ).pack(anchor="w", pady=(12, 0))

    principal = ttk.Frame(corpo, style="Page.TFrame")
    principal.grid(row=0, column=1, rowspan=2, sticky="nsew")
    principal.rowconfigure(1, weight=1)

    card_arquivo = ttk.Frame(principal, style="Section.TFrame", padding=(18, 18))
    card_arquivo.grid(row=0, column=0, sticky="ew")
    ttk.Label(card_arquivo, text="Selecionar arquivo", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        card_arquivo,
        text="Importe um arquivo local e deixe a leitura automatica preparar os produtos.",
        style="SectionText.TLabel",
    ).pack(anchor="w", pady=(4, 12))

    linha_arquivo = ttk.Frame(card_arquivo, style="Section.TFrame")
    linha_arquivo.pack(fill="x")
    entrada_arquivo = ttk.Entry(linha_arquivo, textvariable=arquivo_var)
    entrada_arquivo.pack(side="left", fill="x", expand=True)

    def escolher_arquivo() -> None:
        caminho = filedialog.askopenfilename(
            title="Selecione a entrada",
            filetypes=[
                ("Excel, PDF e Texto", "*.xlsx *.xlsm *.pdf *.txt *.md"),
                ("Todos os arquivos", "*.*"),
            ],
        )
        if caminho:
            arquivo_var.set(caminho)
            modo_var.set("arquivo")
            status_var.set("Arquivo selecionado. A leitura vai usar o caminho informado.")

    ttk.Button(linha_arquivo, text="Procurar", command=escolher_arquivo, style="Accent.TButton").pack(side="left", padx=(10, 0))

    card_texto = ttk.Frame(principal, style="Section.TFrame", padding=(18, 18))
    card_texto.grid(row=1, column=0, sticky="nsew", pady=(14, 0))
    card_texto.rowconfigure(1, weight=1)
    ttk.Label(card_texto, text="Colar texto bruto", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        card_texto,
        text="Cole a mensagem como ela veio. Linhas 'Validade' e cabecalhos de conversa serao ignorados.",
        style="SectionText.TLabel",
        wraplength=760,
        justify="left",
    ).pack(anchor="w", pady=(4, 10))
    entrada_texto = scrolledtext.ScrolledText(card_texto, wrap="word", height=20)
    entrada_texto.pack(fill="both", expand=True)
    entrada_texto.configure(
        font=("Consolas", 11),
        bg="#fffefb",
        fg="#102033",
        insertbackground="#102033",
        relief="flat",
        borderwidth=0,
        highlightthickness=1,
        highlightbackground="#d8d0c4",
        highlightcolor="#0f766e",
        padx=14,
        pady=14,
    )

    rodape = ttk.Frame(root, style="Page.TFrame", padding=(20, 14))
    rodape.pack(fill="x")

    ttk.Label(
        rodape,
        text="Atalhos: Ctrl+Enter continua  •  Esc cancela",
        style="Muted.TLabel",
    ).pack(side="left")

    def atualizar_estado_interface(*_args) -> None:
        modo = modo_var.get()
        arquivo_ativo = modo == "arquivo"
        if arquivo_ativo:
            caminho_atual = normalizar_texto(arquivo_var.get())
            if caminho_atual:
                status_var.set("Modo arquivo ativo. Entrada pronta para leitura.")
            else:
                status_var.set("Modo arquivo ativo. Selecione a planilha, PDF ou TXT para continuar.")
            entrada_arquivo.state(["!disabled"])
            entrada_texto.configure(state="disabled", bg="#f1ede6")
        else:
            status_var.set("Modo texto bruto ativo. Cole a mensagem completa da oferta.")
            entrada_arquivo.state(["disabled"])
            entrada_texto.configure(state="normal", bg="#fffefb")

    def confirmar() -> None:
        modo = modo_var.get()
        caminho = normalizar_texto(arquivo_var.get())
        texto = entrada_texto.get("1.0", tk.END).strip()
        if modo == "arquivo":
            if not caminho:
                messagebox.showerror("Entrada obrigatoria", "Selecione um arquivo para continuar.")
                return
            resultado["tipo"] = "arquivo"
            resultado["valor"] = caminho
            root.destroy()
            return
        if not texto:
            messagebox.showerror("Texto obrigatorio", "Cole o texto bruto para continuar.")
            return
        resultado["tipo"] = "texto"
        resultado["valor"] = texto
        root.destroy()

    def cancelar() -> None:
        if messagebox.askyesno("Cancelar", "Cancelar a producao?"):
            root.destroy()

    ttk.Button(rodape, text="Cancelar", command=cancelar, style="Soft.TButton").pack(side="right")
    ttk.Button(rodape, text="Continuar", command=confirmar, style="Accent.TButton").pack(side="right", padx=(0, 10))
    modo_var.trace_add("write", atualizar_estado_interface)
    arquivo_var.trace_add("write", atualizar_estado_interface)
    atualizar_estado_interface()
    root.bind("<Control-Return>", lambda _event: confirmar())
    root.bind("<Escape>", lambda _event: cancelar())
    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()

    if not resultado:
        return None
    return resultado


def construir_saida_default(arquivo_cdr: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return arquivo_cdr.with_name(f"{arquivo_cdr.stem}_atualizado_{stamp}{arquivo_cdr.suffix}")


def carregar_pagina_configuracao_web() -> str:
    caminho = Path(__file__).with_name(WEB_CONFIG_PAGE_FILE)
    return caminho.read_text(encoding="utf-8")


def remover_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", texto) if not unicodedata.combining(c)
    )


def normalizar_chave_palavra(texto: str) -> str:
    base = remover_acentos(normalizar_texto(texto).lower())
    return re.sub(r"[^a-z]", "", base)


def normalizar_espacamento_produto(texto: str) -> str:
    txt = normalizar_texto(texto).replace("_", " ")
    txt = re.sub(r"([A-Za-zÀ-ÿ])(\d)", r"\1 \2", txt)
    txt = re.sub(r"(\d)([A-Za-zÀ-ÿ])", r"\1 \2", txt)
    txt = re.sub(r"\s*/\s*", "/", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def limpar_texto_produto_para_matching(texto: str) -> str:
    txt = normalizar_espacamento_produto(texto)
    txt = re.sub(r"\bvalidade\b.*$", " ", txt, flags=re.IGNORECASE)
    txt = DATA_RE.sub(" ", txt)
    txt = DATA_SOLTA_RE.sub(" ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def _formatar_numero_canonico(valor: float) -> str:
    if abs(valor - round(valor)) < 0.0001:
        return str(int(round(valor)))
    return f"{valor:.3f}".rstrip("0").rstrip(".")


def normalizar_medida_equivalente(numero: str, unidade: str) -> Optional[str]:
    try:
        valor = float(str(numero).replace(",", "."))
    except Exception:
        return None

    unidade_limpa = normalizar_chave_palavra(unidade)
    if not unidade_limpa:
        return None

    if unidade_limpa in {"kg", "quilo", "quilos", "kilo", "kilos", "kilograma", "kilogramas"}:
        return f"massa:{_formatar_numero_canonico(valor * 1000.0)}g"
    if unidade_limpa in {"g", "gr", "grama", "gramas"}:
        return f"massa:{_formatar_numero_canonico(valor)}g"
    if unidade_limpa == "mg":
        return f"massa:{_formatar_numero_canonico(valor / 1000.0)}g"
    if unidade_limpa in {"l", "lt", "litro", "litros"}:
        return f"volume:{_formatar_numero_canonico(valor * 1000.0)}ml"
    if unidade_limpa == "ml":
        return f"volume:{_formatar_numero_canonico(valor)}ml"
    if unidade_limpa in {"un", "und", "unid", "unidade", "unidades"}:
        return f"contagem:{_formatar_numero_canonico(valor)}un"
    return None


def extrair_medidas_equivalentes(texto: str) -> List[str]:
    medidas: List[str] = []
    vistos = set()
    for match in MEDIDA_RE.finditer(limpar_texto_produto_para_matching(texto)):
        assinatura = normalizar_medida_equivalente(
            match.group("numero"),
            match.group("unidade"),
        )
        if not assinatura or assinatura in vistos:
            continue
        vistos.add(assinatura)
        medidas.append(assinatura)
    return medidas


def base_produto_para_matching(texto: str) -> str:
    txt = limpar_texto_produto_para_matching(texto)
    txt = MEDIDA_RE.sub(" ", txt)
    txt = remover_acentos(txt.lower())
    txt = re.sub(r"[^a-z0-9/&+]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def fingerprint_produto(descricao: str) -> Dict[str, object]:
    base = base_produto_para_matching(descricao)
    medidas = extrair_medidas_equivalentes(descricao)
    tokens = [token for token in base.split() if token]
    return {
        "base": base,
        "tokens": tokens,
        "medidas": medidas,
    }


def similaridade_tokens(a: Sequence[str], b: Sequence[str]) -> float:
    set_a = {item for item in a if item}
    set_b = {item for item in b if item}
    if not set_a and not set_b:
        return 1.0
    if not set_a or not set_b:
        return 0.0
    return len(set_a & set_b) / float(len(set_a | set_b))


def comparar_medidas(
    medidas_a: Sequence[str],
    medidas_b: Sequence[str],
) -> Tuple[float, bool]:
    set_a = {item for item in medidas_a if item}
    set_b = {item for item in medidas_b if item}
    if not set_a and not set_b:
        return 0.60, False
    if set_a == set_b:
        return 1.0, False
    if not set_a or not set_b:
        return 0.68, False

    mapa_a = {item.split(":", 1)[0]: item for item in set_a if ":" in item}
    mapa_b = {item.split(":", 1)[0]: item for item in set_b if ":" in item}
    categorias_compartilhadas = set(mapa_a) & set(mapa_b)
    conflitos = 0
    correspondencias = 0
    for categoria in categorias_compartilhadas:
        if mapa_a[categoria] == mapa_b[categoria]:
            correspondencias += 1
        else:
            conflitos += 1

    if conflitos > 0 and correspondencias == 0:
        return 0.0, True

    intersecao = len(set_a & set_b)
    uniao = max(1, len(set_a | set_b))
    if intersecao > 0:
        return 0.72 + (0.28 * (intersecao / float(uniao))), False
    return 0.30, False


def descricao_parece_produto(descricao: str) -> bool:
    fp = fingerprint_produto(descricao)
    base = str(fp.get("base", ""))
    if not base:
        return False
    letras = re.sub(r"[^a-z]", "", base)
    return len(letras) >= 3


def chave_produto_legacy(descricao: str) -> str:
    base = remover_acentos(normalizar_texto(descricao).lower())
    base = re.sub(r"[^a-z0-9]+", " ", base).strip()
    base = re.sub(r"\s+", " ", base)
    return base


def carregar_corretor_acentos(caminho: Path) -> Dict[str, str]:
    if not caminho.exists():
        return {}

    try:
        with caminho.open("r", encoding="utf-8-sig") as f:
            dados = json.load(f)
    except Exception:
        return {}

    if not isinstance(dados, dict):
        return {}

    bruto = dados.get("palavras")
    if not isinstance(bruto, dict):
        bruto = {k: v for k, v in dados.items() if isinstance(k, str) and isinstance(v, str)}

    saida: Dict[str, str] = {}
    for k, v in bruto.items():
        if not isinstance(k, str) or not isinstance(v, str):
            continue
        chave = normalizar_chave_palavra(k)
        palavra = normalizar_texto(v)
        if not chave or not palavra:
            continue
        saida[chave] = palavra
    return saida


def aplicar_caixa_palavra(origem: str, destino: str) -> str:
    if not origem:
        return destino
    if origem.isupper():
        return destino.upper()
    if origem[:1].isupper() and origem[1:].islower():
        return destino[:1].upper() + destino[1:]
    return destino


def corrigir_acentos_texto(texto: str, corretor: Dict[str, str]) -> str:
    txt = normalizar_texto(texto)
    if not txt or not corretor:
        return txt

    def _trocar(match: re.Match[str]) -> str:
        palavra = match.group(0)
        chave = normalizar_chave_palavra(palavra)
        if not chave:
            return palavra
        corrigida = corretor.get(chave)
        if not corrigida:
            return palavra
        return aplicar_caixa_palavra(palavra, corrigida)

    return WORD_TOKEN_RE.sub(_trocar, txt)


def aplicar_corretor_acentos_produtos(
    produtos: List[Dict[str, str]],
    corretor: Dict[str, str],
) -> Tuple[List[Dict[str, str]], int]:
    if not corretor:
        return produtos, 0

    saida: List[Dict[str, str]] = []
    alterados = 0
    for p in produtos:
        novo = dict(p)
        descricao = normalizar_texto(novo.get("descricao", ""))
        descricao_corrigida = corrigir_acentos_texto(descricao, corretor)
        if descricao_corrigida != descricao:
            alterados += 1
        novo["descricao"] = descricao_corrigida
        saida.append(novo)
    return saida, alterados


def chave_produto(descricao: str) -> str:
    fp = fingerprint_produto(descricao)
    base = str(fp.get("base", "")).strip()
    medidas = [str(item).strip() for item in fp.get("medidas", []) if str(item).strip()]
    partes = [base] + medidas
    return " | ".join(parte for parte in partes if parte)


def normalizar_unidade_saida(valor: str) -> str:
    txt = remover_acentos(normalizar_texto(valor).lower()).replace(".", "")
    if txt in ("kg", "quilo", "quilos", "kilograma", "kilogramas"):
        return "Kg"
    if txt in ("pct", "pcto", "pacote", "pacotes"):
        return "PCT."
    if txt in ("pack", "multipack", "kit", "combo"):
        return "PACK."
    if txt in ("bdj", "bandeja", "bandejas"):
        return "BDJ."
    return "Unid."


def carregar_aprendizado(caminho: Path) -> Dict[str, Dict[str, object]]:
    if not caminho.exists():
        return {}
    try:
        with caminho.open("r", encoding="utf-8-sig") as f:
            dados = json.load(f)
    except Exception:
        return {}

    if not isinstance(dados, dict):
        return {}
    itens = dados.get("itens")
    if not isinstance(itens, dict):
        return {}

    saida: Dict[str, Dict[str, object]] = {}
    for k, v in itens.items():
        if not isinstance(k, str) or not isinstance(v, dict):
            continue
        saida[k] = v
    return saida


def salvar_aprendizado(caminho: Path, itens: Dict[str, Dict[str, object]]) -> None:
    payload = {
        "versao": 2,
        "atualizado_em": dt.datetime.now().isoformat(timespec="seconds"),
        "itens": itens,
    }
    with caminho.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def carregar_aprendizado_github(
    token: Optional[str],
    repo: str,
    caminho_arquivo: str,
    branch: str,
) -> Dict[str, Dict[str, object]]:
    parsed, _sha = carregar_json_arquivo_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
    )
    if not isinstance(parsed, dict):
        return {}
    itens = parsed.get("itens")
    if not isinstance(itens, dict):
        return {}

    saida: Dict[str, Dict[str, object]] = {}
    for k, v in itens.items():
        if not isinstance(k, str) or not isinstance(v, dict):
            continue
        saida[k] = v
    return saida


def salvar_aprendizado_github(
    token: str,
    repo: str,
    caminho_arquivo: str,
    branch: str,
    itens: Dict[str, Dict[str, object]],
) -> None:
    payload = {
        "versao": 2,
        "atualizado_em": dt.datetime.now().isoformat(timespec="seconds"),
        "itens": itens,
    }
    atual, sha_atual = carregar_json_arquivo_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
    )
    if atual is not None and not isinstance(atual, dict):
        raise RuntimeError("O arquivo remoto de aprendizado nao esta em formato JSON de objeto.")
    salvar_arquivo_json_github(
        token=token,
        repo=repo,
        caminho_arquivo=caminho_arquivo,
        branch=branch,
        conteudo_json=payload,
        sha_atual=sha_atual,
        mensagem_commit="Atualiza aprendizado de produtos das ofertas em validade",
    )


def chave_template_execucao(caminho: Path) -> str:
    try:
        return str(caminho.resolve()).strip().lower()
    except Exception:
        return str(caminho).strip().lower()


def carregar_perfil_velocidade(caminho: Path) -> Dict[str, Dict[str, object]]:
    if not caminho.exists():
        return {}
    try:
        with caminho.open("r", encoding="utf-8-sig") as f:
            dados = json.load(f)
    except Exception:
        return {}

    if not isinstance(dados, dict):
        return {}

    templates = dados.get("templates")
    if not isinstance(templates, dict):
        return {}

    saida: Dict[str, Dict[str, object]] = {}
    for k, v in templates.items():
        if isinstance(k, str) and isinstance(v, dict):
            saida[k] = dict(v)
    return saida


def salvar_perfil_velocidade(caminho: Path, perfis: Dict[str, Dict[str, object]]) -> None:
    payload = {
        "versao": 1,
        "atualizado_em": dt.datetime.now().isoformat(timespec="seconds"),
        "templates": perfis,
    }
    with caminho.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def decidir_modo_rapido_inteligente(perfil: Optional[Dict[str, object]]) -> bool:
    if not isinstance(perfil, dict):
        return True

    try:
        falhas_seq = int(perfil.get("falhas_consecutivas_modo_rapido", 0))
    except Exception:
        falhas_seq = 0

    modo_salvo = perfil.get("modo_rapido_ativo")
    if modo_salvo is False:
        # O perfil pode ficar marcado como "false" apos falhas antigas.
        # Se a sequencia de falhas ja foi zerada por execucoes posteriores,
        # permite nova tentativa para reativar a sessao continua.
        return falhas_seq < 2
    if isinstance(modo_salvo, bool):
        return modo_salvo
    return falhas_seq < 2


def registrar_resultado_velocidade(
    perfis: Dict[str, Dict[str, object]],
    template_key: str,
    modo_rapido_ativo: bool,
    tempos_placa: List[float],
    sucesso: bool,
    erro: str = "",
) -> Dict[str, object]:
    perfil_antigo = perfis.get(template_key, {})
    perfil: Dict[str, object] = dict(perfil_antigo if isinstance(perfil_antigo, dict) else {})

    try:
        exec_ok = int(perfil.get("execucoes_ok", 0))
    except Exception:
        exec_ok = 0
    try:
        exec_falha = int(perfil.get("execucoes_falha", 0))
    except Exception:
        exec_falha = 0
    try:
        falhas_seq = int(perfil.get("falhas_consecutivas_modo_rapido", 0))
    except Exception:
        falhas_seq = 0

    if sucesso:
        exec_ok += 1
        falhas_seq = 0
        perfil["execucoes_ok"] = exec_ok
        perfil["falhas_consecutivas_modo_rapido"] = falhas_seq

        if tempos_placa:
            media_execucao = float(sum(tempos_placa) / max(1, len(tempos_placa)))
            try:
                media_historica = float(perfil.get("media_segundos_por_placa", 0.0))
            except Exception:
                media_historica = 0.0
            if media_historica > 0:
                media_nova = (media_historica * 0.65) + (media_execucao * 0.35)
            else:
                media_nova = media_execucao
            perfil["media_segundos_por_placa"] = round(media_nova, 3)
            perfil["ultima_media_segundos_por_placa"] = round(media_execucao, 3)
            perfil["ultimo_total_tentativas"] = int(len(tempos_placa))
            perfil["ultimo_total_segundos"] = round(float(sum(tempos_placa)), 3)
    else:
        exec_falha += 1
        perfil["execucoes_falha"] = exec_falha
        if modo_rapido_ativo:
            falhas_seq += 1
            if falhas_seq >= 2:
                perfil["modo_rapido_ativo"] = False
        perfil["falhas_consecutivas_modo_rapido"] = falhas_seq
        if erro:
            perfil["ultimo_erro"] = normalizar_texto(erro)[:260]

    if sucesso and modo_rapido_ativo:
        perfil["modo_rapido_ativo"] = True
    elif "modo_rapido_ativo" not in perfil:
        perfil["modo_rapido_ativo"] = bool(modo_rapido_ativo)

    perfil["ultima_execucao_em"] = dt.datetime.now().isoformat(timespec="seconds")
    perfis[template_key] = perfil
    return perfil


def registro_aprendizado_valido(chave: str, registro: Dict[str, object]) -> bool:
    descricao = normalizar_texto(str(registro.get("descricao", "")))
    if not descricao or not descricao_parece_produto(descricao):
        return False

    fp_desc = fingerprint_produto(descricao)
    base_desc = str(fp_desc.get("base", ""))
    if not base_desc:
        return False

    chave_limpa = normalizar_texto(chave)
    if not chave_limpa:
        return True

    fp_chave = fingerprint_produto(chave_limpa)
    base_chave = str(fp_chave.get("base", ""))
    if not base_chave:
        return True

    _, conflito_medida = comparar_medidas(
        fp_chave.get("medidas", []),  # type: ignore[arg-type]
        fp_desc.get("medidas", []),  # type: ignore[arg-type]
    )
    if conflito_medida:
        return False

    score_texto = SequenceMatcher(None, base_chave, base_desc).ratio()
    score_tokens = similaridade_tokens(
        fp_chave.get("tokens", []),  # type: ignore[arg-type]
        fp_desc.get("tokens", []),  # type: ignore[arg-type]
    )
    try:
        hits = int(registro.get("hits", 0) or 0)
    except Exception:
        hits = 0
    score_final = max(score_texto, score_tokens)
    return score_final >= 0.34 or (hits >= 4 and score_final >= 0.28)


def aliases_registro_aprendizado(chave: str, registro: Dict[str, object]) -> List[str]:
    aliases: List[str] = []
    vistos = set()
    candidatos = [
        chave,
        chave_produto_legacy(chave),
        str(registro.get("descricao", "") or ""),
        chave_produto(str(registro.get("descricao", "") or "")),
        chave_produto_legacy(str(registro.get("descricao", "") or "")),
    ]
    bruto_aliases = registro.get("aliases")
    if isinstance(bruto_aliases, list):
        candidatos.extend(str(item or "") for item in bruto_aliases)

    for candidato in candidatos:
        for normalizado in (chave_produto(candidato), chave_produto_legacy(candidato)):
            valor = normalizar_texto(normalizado)
            if not valor or valor in vistos:
                continue
            vistos.add(valor)
            aliases.append(valor)
    return aliases


def buscar_aprendizado(
    descricao: str,
    itens: Dict[str, Dict[str, object]],
    similaridade_minima: float = 0.78,
) -> Optional[Dict[str, object]]:
    fp_busca = fingerprint_produto(descricao)
    chave = chave_produto(descricao)
    chave_legacy = chave_produto_legacy(descricao)
    base_busca = str(fp_busca.get("base", ""))
    if not base_busca:
        return None

    for chave_exata in (chave, chave_legacy):
        if chave_exata and chave_exata in itens:
            rec = itens[chave_exata]
            if not isinstance(rec, dict) or not registro_aprendizado_valido(chave_exata, rec):
                continue
            return {
                "descricao": str(rec.get("descricao", descricao)),
                "unidade": normalizar_unidade_saida(str(rec.get("unidade", "Unid."))),
                "fonte": "exato",
                "score": 1.0,
                "chave": chave_exata,
            }

    melhor_key = ""
    melhor_rec: Optional[Dict[str, object]] = None
    melhor_score = 0.0
    for k, rec in itens.items():
        if not isinstance(k, str) or not isinstance(rec, dict):
            continue
        if not registro_aprendizado_valido(k, rec):
            continue

        aliases = aliases_registro_aprendizado(k, rec)
        melhor_alias_score = 0.0
        conflito_em_alias = False
        for alias in aliases:
            fp_alias = fingerprint_produto(alias)
            base_alias = str(fp_alias.get("base", ""))
            if not base_alias:
                continue

            score_base = SequenceMatcher(None, base_busca, base_alias).ratio()
            score_tokens = similaridade_tokens(
                fp_busca.get("tokens", []),  # type: ignore[arg-type]
                fp_alias.get("tokens", []),  # type: ignore[arg-type]
            )
            score_medidas, conflito_medidas = comparar_medidas(
                fp_busca.get("medidas", []),  # type: ignore[arg-type]
                fp_alias.get("medidas", []),  # type: ignore[arg-type]
            )
            if conflito_medidas:
                conflito_em_alias = True
                continue

            score_total = (max(score_base, score_tokens) * 0.60) + (score_tokens * 0.15) + (score_medidas * 0.25)
            if score_total > melhor_alias_score:
                melhor_alias_score = score_total

        if conflito_em_alias and melhor_alias_score <= 0.0:
            continue
        if melhor_alias_score > melhor_score:
            melhor_score = melhor_alias_score
            melhor_key = k
            melhor_rec = rec

    if not melhor_key or melhor_rec is None or melhor_score < similaridade_minima:
        return None

    return {
        "descricao": str(melhor_rec.get("descricao", descricao)),
        "unidade": normalizar_unidade_saida(str(melhor_rec.get("unidade", "Unid."))),
        "fonte": "similar",
        "score": melhor_score,
        "chave": melhor_key,
    }


def aplicar_aprendizado(
    produtos: List[Dict[str, str]],
    itens_aprendidos: Dict[str, Dict[str, object]],
) -> Tuple[List[Dict[str, str]], int]:
    if not itens_aprendidos:
        return produtos, 0

    novos: List[Dict[str, str]] = []
    usados = 0
    for p in produtos:
        desc = str(p.get("descricao", ""))
        sugestao = buscar_aprendizado(desc, itens_aprendidos)
        if sugestao is None:
            novos.append(dict(p))
            continue

        novo = dict(p)
        novo["descricao"] = normalizar_texto(str(sugestao["descricao"])) or normalizar_texto(desc)
        novo["unidade"] = normalizar_unidade_saida(str(sugestao["unidade"]))
        novos.append(novo)
        usados += 1

    return novos, usados


def registrar_aprendizado(
    originais: List[Dict[str, str]],
    revisados: List[Dict[str, str]],
    itens_aprendidos: Dict[str, Dict[str, object]],
) -> int:
    alterados = 0
    for p_orig, p_rev in zip(originais, revisados):
        desc_orig = normalizar_texto(str(p_orig.get("descricao", "")))
        desc_rev = normalizar_texto(str(p_rev.get("descricao", "")))
        unid_orig = normalizar_unidade_saida(str(p_orig.get("unidade", "Unid.")))
        unid_rev = normalizar_unidade_saida(str(p_rev.get("unidade", "Unid.")))

        if not desc_orig or not desc_rev:
            continue
        if desc_orig == desc_rev and unid_orig == unid_rev:
            continue

        chave = chave_produto(desc_orig)
        if not chave or not descricao_parece_produto(desc_rev):
            continue

        rec = itens_aprendidos.get(chave, {})
        hits_antigos = int(rec.get("hits", 0)) if isinstance(rec, dict) else 0
        aliases = set()
        if isinstance(rec, dict):
            bruto_aliases = rec.get("aliases")
            if isinstance(bruto_aliases, list):
                aliases.update(normalizar_texto(str(item or "")) for item in bruto_aliases if normalizar_texto(str(item or "")))
        aliases.update(
            item
            for item in {
                chave,
                chave_produto_legacy(desc_orig),
                chave_produto(desc_rev),
                chave_produto_legacy(desc_rev),
            }
            if item
        )
        fp_rev = fingerprint_produto(desc_rev)
        itens_aprendidos[chave] = {
            "descricao": desc_rev,
            "unidade": unid_rev,
            "hits": hits_antigos + 1,
            "aliases": sorted(aliases),
            "descricao_base": str(fp_rev.get("base", "")),
            "medidas": list(fp_rev.get("medidas", [])),
            "confidence": 1.0,
            "atualizado_em": dt.datetime.now().isoformat(timespec="seconds"),
        }
        alterados += 1

    return alterados


def encontrar_executavel_ollama() -> Optional[Path]:
    global _OLLAMA_EXECUTABLE_CACHE, _OLLAMA_LOOKUP_DONE
    if _OLLAMA_LOOKUP_DONE:
        return _OLLAMA_EXECUTABLE_CACHE
    _OLLAMA_LOOKUP_DONE = True

    caminho_path = shutil.which("ollama")
    if caminho_path:
        _OLLAMA_EXECUTABLE_CACHE = Path(caminho_path)
        return _OLLAMA_EXECUTABLE_CACHE

    candidatos: List[Path] = []
    for var_ambiente in ("LOCALAPPDATA", "ProgramFiles", "ProgramFiles(x86)"):
        base = os.environ.get(var_ambiente, "").strip()
        if not base:
            continue
        candidatos.extend(
            [
                Path(base) / "Programs" / "Ollama" / "ollama.exe",
                Path(base) / "Ollama" / "ollama.exe",
            ]
        )

    for candidato in candidatos:
        try:
            if candidato.exists():
                _OLLAMA_EXECUTABLE_CACHE = candidato
                return _OLLAMA_EXECUTABLE_CACHE
        except Exception:
            continue
    return None


def extrair_json_objeto(texto: str) -> Optional[Dict[str, Any]]:
    bruto = normalizar_texto(texto)
    if not bruto:
        return None
    candidatos = [bruto]
    candidatos.extend(re.findall(r"\{[\s\S]*\}", bruto))
    for candidato in candidatos:
        try:
            parsed = json.loads(candidato)
        except Exception:
            continue
        if isinstance(parsed, dict):
            return parsed
    return None


def descricao_precisa_ia(texto: str) -> bool:
    limpa = normalizar_espacamento_produto(texto)
    if not limpa:
        return False
    compacta = re.sub(r"[^a-z0-9]+", "", remover_acentos(limpa.lower()))
    if len(compacta) <= 4:
        return False
    if re.search(r"[�?]", limpa):
        return True
    if re.search(r"\d{2,}\s*[xX]\s*\d{2,}", limpa):
        return True
    if " " not in limpa and len(compacta) >= 12:
        return True
    if re.search(r"[A-Za-z][0-9][A-Za-z]", limpa):
        return True
    if len(re.findall(r"[A-Za-z]", limpa)) >= 8 and limpa == limpa.lower():
        return True
    return False


def chamar_ia_local_limpeza_nome(
    descricao: str,
    unidade_atual: str,
    modelo: str,
    timeout_segundos: float,
) -> Optional[Dict[str, Any]]:
    global _OLLAMA_WARNING_EMITIDO
    executavel = encontrar_executavel_ollama()
    if not executavel:
        if not _OLLAMA_WARNING_EMITIDO:
            print("IA local: Ollama nao encontrado. Seguindo apenas com heuristicas e aprendizado local.")
            _OLLAMA_WARNING_EMITIDO = True
        return None

    prompt = (
        "Corrija o nome de um produto para uso em uma placa de oferta. "
        "Responda somente JSON valido, sem markdown, com as chaves "
        "corrected_name, unit, confidence. "
        "Regras: nao invente informacoes; preserve marca e gramatura/volume reais; "
        "corrija OCR, acentos, espacos e abreviacoes quando fizer sentido comercial; "
        "unit deve ser apenas 'Kg', 'Unid.', 'PCT.', 'PACK.' ou 'BDJ.'; "
        "se houver peso/volume de embalagem como 500g, 1kg, 250ml, 2L, prefira 'Unid.'; "
        "se indicar pacote, prefira 'PCT.'; se indicar pack, kit ou combo, prefira 'PACK.'; "
        "se indicar bandeja ou ovos, prefira 'BDJ.'; "
        "se for item vendido por peso a granel, use 'Kg'.\n"
        f'Descricao: "{descricao}"\n'
        f'Unidade atual: "{unidade_atual}"'
    )
    try:
        proc = subprocess.run(
            [str(executavel), "run", str(modelo or DEFAULT_OLLAMA_MODEL), prompt],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=max(5, int(timeout_segundos)),
            check=False,
        )
    except Exception:
        return None

    if int(proc.returncode) != 0:
        stderr = normalizar_texto(proc.stderr or "")
        if stderr and not _OLLAMA_WARNING_EMITIDO:
            print(f"IA local: Ollama indisponivel no momento ({stderr}).")
            _OLLAMA_WARNING_EMITIDO = True
        return None
    return extrair_json_objeto(proc.stdout)


def aplicar_limpeza_ia_local(
    produtos: List[Dict[str, str]],
    itens_aprendidos: Dict[str, Dict[str, object]],
    limite_itens: int,
    modelo: str,
    timeout_segundos: float,
) -> Tuple[List[Dict[str, str]], int]:
    if limite_itens <= 0:
        return produtos, 0

    saida: List[Dict[str, str]] = []
    usados = 0
    tentativas = 0
    for produto in produtos:
        novo = dict(produto)
        descricao = normalizar_texto(novo.get("descricao", ""))
        unidade = normalizar_unidade_saida(str(novo.get("unidade", "Unid.")))
        if (
            tentativas < limite_itens
            and descricao
            and descricao_precisa_ia(descricao)
            and buscar_aprendizado(descricao, itens_aprendidos) is None
        ):
            tentativas += 1
            resposta = chamar_ia_local_limpeza_nome(
                descricao=descricao,
                unidade_atual=unidade,
                modelo=modelo,
                timeout_segundos=timeout_segundos,
            )
            if isinstance(resposta, dict):
                nome_corrigido = normalizar_texto(str(resposta.get("corrected_name", "")))
                unidade_corrigida = normalizar_unidade_saida(str(resposta.get("unit", unidade)))
                try:
                    confianca = float(resposta.get("confidence", 0.0) or 0.0)
                except Exception:
                    confianca = 0.0
                if nome_corrigido and confianca >= 0.55:
                    novo["descricao"] = nome_corrigido
                    novo["unidade"] = unidade_corrigida
                    usados += 1
        saida.append(novo)
    return saida, usados


def processar_produtos_inteligentes(
    produtos: List[Dict[str, str]],
    corretor_acentos: Dict[str, str],
    itens_aprendidos: Dict[str, Dict[str, object]],
    usar_aprendizado: bool,
    usar_ia_local: bool,
    limite_itens_ia_local: int,
    modelo_ia_local: str,
    timeout_ia_local: float,
) -> Tuple[List[Dict[str, str]], Dict[str, int]]:
    stats = {
        "acentos": 0,
        "aprendizado": 0,
        "ia_local": 0,
        "acentos_pos_ia": 0,
        "aprendizado_pos_ia": 0,
    }

    processados = [dict(item) for item in produtos]

    if corretor_acentos:
        processados, stats["acentos"] = aplicar_corretor_acentos_produtos(processados, corretor_acentos)

    if usar_aprendizado and itens_aprendidos:
        processados, stats["aprendizado"] = aplicar_aprendizado(processados, itens_aprendidos)

    if usar_ia_local:
        processados, stats["ia_local"] = aplicar_limpeza_ia_local(
            produtos=processados,
            itens_aprendidos=itens_aprendidos,
            limite_itens=max(0, int(limite_itens_ia_local)),
            modelo=str(modelo_ia_local or DEFAULT_OLLAMA_MODEL),
            timeout_segundos=max(1.0, float(timeout_ia_local)),
        )
        if corretor_acentos:
            processados, stats["acentos_pos_ia"] = aplicar_corretor_acentos_produtos(processados, corretor_acentos)
        if usar_aprendizado and itens_aprendidos:
            processados, stats["aprendizado_pos_ia"] = aplicar_aprendizado(processados, itens_aprendidos)

    return processados, stats


def analisar_entrada_web_payload(
    payload: Dict[str, Any],
    corretor_acentos: Dict[str, str],
    itens_aprendidos: Dict[str, Dict[str, object]],
    usar_aprendizado: bool,
    usar_ia_local: bool,
    limite_itens_ia_local: int,
    modelo_ia_local: str,
    timeout_ia_local: float,
) -> Dict[str, Any]:
    modo = normalizar_texto(payload.get("mode", "")).lower()
    produtos_brutos: List[Dict[str, str]]
    origem = ""

    if modo == "text":
        texto = str(payload.get("text", "") or "")
        if not normalizar_texto(texto):
            raise EntradaInvalida("Cole o texto bruto antes de analisar.")
        origem = "texto bruto (pagina web)"
        produtos_brutos = extrair_texto_bruto_lista(texto)
    elif modo == "file":
        nome_arquivo = normalizar_texto(payload.get("file_name", "entrada"))
        conteudo_b64 = str(payload.get("file_content_b64", "") or "")
        if not conteudo_b64:
            raise EntradaInvalida("Selecione um arquivo antes de analisar.")
        suffix = Path(nome_arquivo or "entrada.tmp").suffix or ".tmp"
        try:
            conteudo = base64.b64decode(conteudo_b64, validate=True)
        except Exception as exc:
            raise EntradaInvalida("Nao foi possivel decodificar o arquivo enviado.") from exc

        temp_path: Optional[Path] = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
                temp_file.write(conteudo)
                temp_path = Path(temp_file.name)
            origem = nome_arquivo or str(temp_path.name)
            produtos_brutos = extrair_produtos(temp_path)
        finally:
            if temp_path is not None:
                try:
                    temp_path.unlink(missing_ok=True)
                except Exception:
                    pass
    else:
        raise EntradaInvalida("Escolha Arquivo ou Texto bruto antes de analisar.")

    produtos_processados, stats = processar_produtos_inteligentes(
        produtos=produtos_brutos,
        corretor_acentos=corretor_acentos,
        itens_aprendidos=itens_aprendidos,
        usar_aprendizado=usar_aprendizado,
        usar_ia_local=usar_ia_local,
        limite_itens_ia_local=limite_itens_ia_local,
        modelo_ia_local=modelo_ia_local,
        timeout_ia_local=timeout_ia_local,
    )

    return {
        "origin": origem,
        "raw_products": produtos_brutos,
        "products": produtos_processados,
        "stats": stats,
        "plate_summary": montar_placas_para_interface(produtos_processados),
    }


def montar_lotes(produtos: List[Dict[str, str]], tamanho_lote: int = 4) -> List[List[Dict[str, str]]]:
    lotes: List[List[Dict[str, str]]] = []
    for i in range(0, len(produtos), tamanho_lote):
        lote = produtos[i : i + tamanho_lote]
        while len(lote) < tamanho_lote:
            lote.append(
                {
                    "descricao": "",
                    "unidade": "",
                    "preco": "",
                    "validade_oferta": "",
                    "codigo_barras": "",
                    "usar_codigo_barras": False,
                }
            )
        lotes.append(lote)
    return lotes


def revisar_produtos_tela(
    produtos: List[Dict[str, str]],
    permitir_selecao_placas: bool = True,
    desligar_ao_final_inicial: bool = False,
) -> Optional[Tuple[List[Dict[str, str]], Optional[List[int]], bool]]:
    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except Exception:
        print("Aviso: tkinter indisponivel. Seguindo sem tela de revisao.")
        return produtos, None, bool(desligar_ao_final_inicial)

    resultado: Dict[str, object] = {
        "ok": False,
        "produtos": produtos,
        "selecionadas": None,
        "desligar_ao_final": bool(desligar_ao_final_inicial),
    }

    root = tk.Tk()
    root.title("Selecao de Placas e Ajustes")
    root.geometry("1460x860")
    root.minsize(1180, 700)
    configurar_estilo_interface(root, ttk)

    cabecalho = ttk.Frame(root, style="Hero.TFrame", padding=(18, 18))
    cabecalho.pack(fill="x")
    ttk.Label(cabecalho, text="Selecao de placas e ajustes", style="HeroTitle.TLabel").pack(anchor="w")
    texto_topo = (
        "Revise os itens carregados da planilha, ajuste descricao, preco, unidade e codigo de barras, "
        "e confirme apenas as placas que devem ser produzidas."
    )
    if permitir_selecao_placas:
        texto_topo += " A selecao permanece sincronizada com os ajustes do lote."
    ttk.Label(
        cabecalho,
        text=texto_topo,
        style="HeroText.TLabel",
        wraplength=1320,
        justify="left",
    ).pack(anchor="w", pady=(6, 0))

    faixa_resumo = ttk.Frame(root, style="Page.TFrame", padding=(16, 12, 16, 4))
    faixa_resumo.pack(fill="x")

    def criar_indicador(parent, titulo: str) -> object:
        card = ttk.Frame(parent, style="Card.TFrame", padding=(14, 10))
        card.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Label(card, text=titulo, style="CardTitle.TLabel").pack(anchor="w")
        var = tk.StringVar(value="0")
        ttk.Label(card, textvariable=var, style="CardValue.TLabel").pack(anchor="w", pady=(4, 0))
        return var

    var_total_itens = criar_indicador(faixa_resumo, "Itens")
    var_itens_validos = criar_indicador(faixa_resumo, "Validos")
    var_total_placas = criar_indicador(faixa_resumo, "Placas")
    var_placas_selecionadas = criar_indicador(faixa_resumo, "Selecionadas")
    var_placas_vazias = criar_indicador(faixa_resumo, "Vazias")

    acoes = ttk.Frame(root, style="Page.TFrame", padding=(16, 6, 16, 8))
    acoes.pack(fill="x")
    var_status_produtos = tk.StringVar(value="")
    ttk.Label(acoes, textvariable=var_status_produtos, style="Status.TLabel").pack(side="right")

    painel = ttk.Panedwindow(root, orient="horizontal")
    painel.pack(fill="both", expand=True, padx=16, pady=(4, 0))

    area_produtos = ttk.Frame(painel, style="Section.TFrame", padding=(14, 14))
    painel.add(area_produtos, weight=5)

    area_placas = None
    if permitir_selecao_placas:
        area_placas = ttk.Frame(painel, style="Section.TFrame", padding=(14, 14))
        painel.add(area_placas, weight=3)

    ttk.Label(area_produtos, text="Ajustes de produtos", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        area_produtos,
        text="Edite os dados do lote antes da producao. As acoes rapidas ajudam a padronizar a unidade.",
        style="SectionText.TLabel",
        wraplength=760,
        justify="left",
    ).pack(anchor="w", pady=(4, 10))

    container = ttk.Frame(area_produtos, style="Section.TFrame")
    container.pack(fill="both", expand=True)

    canvas = tk.Canvas(container, highlightthickness=0)
    scroll_y = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    conteudo = ttk.Frame(canvas)

    conteudo.bind(
        "<Configure>",
        lambda _e: canvas.configure(scrollregion=canvas.bbox("all")),
    )
    canvas_window = canvas.create_window((0, 0), window=conteudo, anchor="nw")
    canvas.configure(yscrollcommand=scroll_y.set)

    def ajustar_largura(_event=None):
        largura = max(260, canvas.winfo_width() - 4)
        canvas.itemconfig(canvas_window, width=largura)

    canvas.bind("<Configure>", ajustar_largura)
    canvas.pack(side="left", fill="both", expand=True)
    scroll_y.pack(side="right", fill="y")

    ttk.Label(conteudo, text="#", width=4).grid(row=0, column=0, sticky="w", padx=(0, 6), pady=(0, 6))
    ttk.Label(conteudo, text="Descricao (editavel)", width=62).grid(
        row=0, column=1, sticky="w", padx=(0, 6), pady=(0, 6)
    )
    ttk.Label(conteudo, text="Preco", width=10).grid(row=0, column=2, sticky="w", padx=(0, 6), pady=(0, 6))
    ttk.Label(conteudo, text="Unidade", width=10).grid(row=0, column=3, sticky="w", pady=(0, 6))
    ttk.Label(conteudo, text="Cod. barras", width=16).grid(row=0, column=4, sticky="w", padx=(6, 6), pady=(0, 6))
    ttk.Label(conteudo, text="Usar CB", width=10).grid(row=0, column=5, sticky="w", padx=(0, 6), pady=(0, 6))
    ttk.Label(conteudo, text="Acao", width=8).grid(row=0, column=6, sticky="w", padx=(8, 0), pady=(0, 6))

    conteudo.grid_columnconfigure(1, weight=1)

    linhas_base: List[Dict[str, str]] = []
    for p in produtos:
        linhas_base.append(
            {
                "descricao": normalizar_texto(p.get("descricao", "")),
                "unidade": normalizar_unidade_saida(str(p.get("unidade", "Unid."))),
                "preco": normalizar_preco_str(str(p.get("preco", ""))),
                "codigo_barras": normalizar_codigo_barras_saida(p.get("codigo_barras", "")),
                "usar_codigo_barras": normalizar_flag_codigo_barras_saida(
                    p.get("usar_codigo_barras", False),
                    p.get("codigo_barras", ""),
                ),
            }
        )

    vars_descricao: List[object] = []
    vars_preco: List[object] = []
    vars_unidade: List[object] = []
    vars_codigo_barras: List[object] = []
    vars_uso_codigo_barras: List[object] = []
    entradas_desc: List[object] = []
    widgets_linhas: List[List[object]] = []
    recalcular_placas_ref: List[object] = [lambda: None]
    atualizar_dashboard_ref: List[object] = [lambda: None]

    def coletar_linhas_da_tela(validar: bool = False) -> Optional[List[Dict[str, str]]]:
        novos: List[Dict[str, str]] = []
        for idx in range(len(vars_descricao)):
            descricao = normalizar_texto(str(vars_descricao[idx].get()))
            preco = normalizar_preco_str(str(vars_preco[idx].get()))
            unidade = normalizar_unidade_saida(str(vars_unidade[idx].get()))
            codigo_barras_bruto = extrair_digitos_codigo_barras(vars_codigo_barras[idx].get())
            usar_codigo_barras = bool(vars_uso_codigo_barras[idx].get()) and len(codigo_barras_bruto) == 13
            if validar and not descricao:
                messagebox.showerror(
                    "Descricao obrigatoria",
                    f"O produto #{idx + 1} esta com descricao vazia. Ajuste antes de continuar.",
                )
                return None
            if validar and descricao and not preco:
                messagebox.showerror(
                    "Preco obrigatorio",
                    f"O produto #{idx + 1} esta sem preco valido. Use formato como 12,99.",
                )
                return None
            if validar and codigo_barras_bruto and len(codigo_barras_bruto) != 13:
                messagebox.showerror(
                    "Codigo de barras invalido",
                    f"O produto #{idx + 1} precisa ter 13 digitos no codigo de barras.",
                )
                return None
            novos.append(
                {
                    "descricao": descricao,
                    "unidade": unidade,
                    "preco": preco,
                    "codigo_barras": codigo_barras_bruto if len(codigo_barras_bruto) == 13 else "",
                    "usar_codigo_barras": usar_codigo_barras,
                }
            )
        return novos

    def renderizar_linhas(linhas: List[Dict[str, str]]) -> None:
        for grupo in widgets_linhas:
            for w in grupo:
                try:
                    w.destroy()
                except Exception:
                    pass
        widgets_linhas.clear()
        vars_descricao.clear()
        vars_preco.clear()
        vars_unidade.clear()
        vars_codigo_barras.clear()
        vars_uso_codigo_barras.clear()
        entradas_desc.clear()

        for i, p in enumerate(linhas, start=1):
            lbl_idx = ttk.Label(conteudo, text=str(i), width=4)
            lbl_idx.grid(row=i, column=0, sticky="nw", padx=(0, 6), pady=2)

            var_d = tk.StringVar(value=normalizar_texto(p.get("descricao", "")))
            entrada_desc = ttk.Entry(conteudo, textvariable=var_d, width=70)
            entrada_desc.grid(row=i, column=1, sticky="ew", padx=(0, 6), pady=2)

            var_p = tk.StringVar(value=normalizar_preco_str(str(p.get("preco", ""))))
            entrada_preco = ttk.Entry(conteudo, textvariable=var_p, width=10)
            entrada_preco.grid(row=i, column=2, sticky="w", padx=(0, 6), pady=2)

            var_u = tk.StringVar(value=normalizar_unidade_saida(str(p.get("unidade", "Unid."))))
            combo = ttk.Combobox(
                conteudo,
                textvariable=var_u,
                values=UNIT_OPTIONS,
                width=9,
                state="readonly",
            )
            combo.grid(row=i, column=3, sticky="w", pady=2)

            var_cb = tk.StringVar(value=normalizar_codigo_barras_saida(p.get("codigo_barras", "")))
            entrada_cb = ttk.Entry(conteudo, textvariable=var_cb, width=18)
            entrada_cb.grid(row=i, column=4, sticky="w", padx=(6, 6), pady=2)

            var_use_cb = tk.BooleanVar(
                value=normalizar_flag_codigo_barras_saida(
                    p.get("usar_codigo_barras", False),
                    p.get("codigo_barras", ""),
                )
            )
            check_cb = ttk.Checkbutton(conteudo, text="Usar", variable=var_use_cb)
            check_cb.grid(row=i, column=5, sticky="w", padx=(0, 6), pady=2)

            btn_excluir = ttk.Button(conteudo, text="Excluir", width=8, command=lambda idx=i - 1: excluir_linha(idx))
            btn_excluir.grid(row=i, column=6, sticky="w", padx=(8, 0), pady=2)

            vars_descricao.append(var_d)
            vars_preco.append(var_p)
            vars_unidade.append(var_u)
            vars_codigo_barras.append(var_cb)
            vars_uso_codigo_barras.append(var_use_cb)
            entradas_desc.append(entrada_desc)
            widgets_linhas.append([lbl_idx, entrada_desc, entrada_preco, combo, entrada_cb, check_cb, btn_excluir])

            var_d.trace_add("write", lambda *_args: atualizar_dashboard_ref[0]())
            var_p.trace_add("write", lambda *_args: atualizar_dashboard_ref[0]())
            var_u.trace_add("write", lambda *_args: atualizar_dashboard_ref[0]())
            var_cb.trace_add("write", lambda *_args: atualizar_dashboard_ref[0]())
            var_use_cb.trace_add("write", lambda *_args: atualizar_dashboard_ref[0]())

    def obter_produtos_editados(validar: bool = False) -> Optional[List[Dict[str, str]]]:
        novos = coletar_linhas_da_tela(validar=validar)
        if novos is None:
            return None
        linhas_base[:] = novos
        return [dict(p) for p in linhas_base]

    def aplicar_todos(valor: str) -> None:
        for var_u in vars_unidade:
            var_u.set(valor)
        recalcular_placas_ref[0]()

    def inserir_linha() -> None:
        atuais = coletar_linhas_da_tela(validar=False)
        if atuais is None:
            return
        linhas_base[:] = atuais
        linhas_base.append(
            {"descricao": "", "unidade": "Unid.", "preco": "", "codigo_barras": "", "usar_codigo_barras": False}
        )
        renderizar_linhas(linhas_base)
        if entradas_desc:
            try:
                entradas_desc[-1].focus_set()
            except Exception:
                pass
        recalcular_placas_ref[0]()

    def duplicar_ultimo_item() -> None:
        atuais = coletar_linhas_da_tela(validar=False)
        if atuais is None:
            return
        if not atuais:
            linhas_base[:] = [
                {"descricao": "", "unidade": "Unid.", "preco": "", "codigo_barras": "", "usar_codigo_barras": False}
            ]
        else:
            linhas_base[:] = atuais
            linhas_base.append(dict(atuais[-1]))
        renderizar_linhas(linhas_base)
        if entradas_desc:
            try:
                entradas_desc[-1].focus_set()
            except Exception:
                pass
        recalcular_placas_ref[0]()

    def excluir_vazios() -> None:
        atuais = coletar_linhas_da_tela(validar=False)
        if atuais is None:
            return
        filtrados = [
            p
            for p in atuais
            if normalizar_texto(p.get("descricao", "")) or normalizar_preco_str(str(p.get("preco", "")))
        ]
        linhas_base[:] = (
            filtrados
            if filtrados
            else [{"descricao": "", "unidade": "Unid.", "preco": "", "codigo_barras": "", "usar_codigo_barras": False}]
        )
        renderizar_linhas(linhas_base)
        recalcular_placas_ref[0]()

    def excluir_linha(indice: int) -> None:
        atuais = coletar_linhas_da_tela(validar=False)
        if atuais is None:
            return
        if indice < 0 or indice >= len(atuais):
            return
        del atuais[indice]
        linhas_base[:] = atuais
        renderizar_linhas(linhas_base)
        if entradas_desc:
            alvo = min(indice, len(entradas_desc) - 1)
            try:
                entradas_desc[alvo].focus_set()
            except Exception:
                pass
        recalcular_placas_ref[0]()

    ttk.Button(acoes, text="Todos Unid.", command=lambda: aplicar_todos("Unid.")).pack(side="left")
    ttk.Button(acoes, text="Todos Kg", command=lambda: aplicar_todos("Kg")).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Todos PCT.", command=lambda: aplicar_todos("PCT.")).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Todos PACK.", command=lambda: aplicar_todos("PACK.")).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Todos BDJ.", command=lambda: aplicar_todos("BDJ.")).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Inserir Item", command=inserir_linha).pack(side="left", padx=(18, 0))
    ttk.Button(acoes, text="Duplicar Ultimo", command=duplicar_ultimo_item).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Excluir Vazios", command=excluir_vazios).pack(side="left", padx=(8, 0))

    renderizar_linhas(linhas_base)

    selecionadas: set[int] = set()
    placas: List[Dict[str, object]] = []
    placas_mapa: Dict[int, Dict[str, object]] = {}
    exibidas: List[int] = []

    def indices_placas_alvo(somente_selecionadas: bool) -> List[int]:
        atuais = coletar_linhas_da_tela(validar=False) or []
        if not somente_selecionadas:
            return list(range(len(atuais)))
        indices: List[int] = []
        for numero in sorted(selecionadas):
            inicio = max(0, (int(numero) - 1) * 4)
            fim = min(len(atuais), inicio + 4)
            indices.extend(range(inicio, fim))
        return indices

    def aplicar_codigo_barras_em_lote(ativo: bool, somente_selecionadas: bool) -> None:
        atuais = coletar_linhas_da_tela(validar=False)
        if atuais is None:
            return
        indices = indices_placas_alvo(somente_selecionadas)
        if somente_selecionadas and not indices:
            return
        for idx in indices:
            if 0 <= idx < len(atuais):
                codigo = normalizar_codigo_barras_saida(atuais[idx].get("codigo_barras", ""))
                atuais[idx]["usar_codigo_barras"] = bool(ativo and codigo)
        linhas_base[:] = atuais
        renderizar_linhas(linhas_base)
        recalcular_placas_ref[0]()

    ttk.Button(acoes, text="CB Todos On", command=lambda: aplicar_codigo_barras_em_lote(True, False)).pack(
        side="left", padx=(18, 0)
    )
    ttk.Button(acoes, text="CB Todos Off", command=lambda: aplicar_codigo_barras_em_lote(False, False)).pack(
        side="left", padx=(8, 0)
    )
    ttk.Button(acoes, text="CB Sel. On", command=lambda: aplicar_codigo_barras_em_lote(True, True)).pack(
        side="left", padx=(18, 0)
    )
    ttk.Button(acoes, text="CB Sel. Off", command=lambda: aplicar_codigo_barras_em_lote(False, True)).pack(
        side="left", padx=(8, 0)
    )

    def atualizar_dashboard() -> None:
        produtos_atuais = coletar_linhas_da_tela(validar=False) or []
        placas_atuais = montar_placas_para_interface(produtos_atuais)
        numeros_validos = {int(p.get("numero", 0)) for p in placas_atuais}
        com_conteudo = sum(1 for p in placas_atuais if int(p.get("qtd_validos", 0)) > 0)
        selecionadas_validas = len([n for n in selecionadas if n in numeros_validos])
        var_total_itens.set(str(len(produtos_atuais)))
        var_itens_validos.set(str(contar_produtos_validos(produtos_atuais)))
        var_total_placas.set(str(len(placas_atuais)))
        var_placas_selecionadas.set(str(selecionadas_validas if permitir_selecao_placas else len(placas_atuais)))
        var_placas_vazias.set(str(max(0, len(placas_atuais) - com_conteudo)))
        vazios = max(0, len(produtos_atuais) - contar_produtos_validos(produtos_atuais))
        var_status_produtos.set(
            f"Itens incompletos: {vazios} | Faixas selecionadas: {resumir_intervalos_numericos(selecionadas)}"
        )

    atualizar_dashboard_ref[0] = atualizar_dashboard

    if area_placas is not None:
        ttk.Label(area_placas, text="Placas para producao", style="SectionTitle.TLabel").pack(anchor="w")
        ttk.Label(
            area_placas,
            text="Use o filtro, confira o preview do lote e escolha apenas o que deve ser impresso.",
            style="SectionText.TLabel",
            wraplength=480,
            justify="left",
        ).pack(anchor="w", pady=(4, 10))

        filtros = ttk.Frame(area_placas, style="Section.TFrame")
        filtros.pack(fill="x")
        ttk.Label(filtros, text="Filtro placas:").pack(side="left")
        var_filtro = tk.StringVar(value="")
        entrada_filtro = ttk.Entry(filtros, textvariable=var_filtro, width=28)
        entrada_filtro.pack(side="left", padx=(8, 8))
        ttk.Label(filtros, text="Status:").pack(side="left")
        var_status_filtro = tk.StringVar(value="Todas")
        combo_status = ttk.Combobox(
            filtros,
            textvariable=var_status_filtro,
            values=("Todas", "Com conteudo", "Vazias", "Selecionadas", "Nao selecionadas"),
            width=17,
            state="readonly",
        )
        combo_status.pack(side="left", padx=(8, 8))
        var_contagem = tk.StringVar(value="")
        ttk.Label(filtros, textvariable=var_contagem, style="Status.TLabel").pack(side="left")

        acoes_placas = ttk.Frame(area_placas, style="Section.TFrame", padding=(0, 8, 0, 8))
        acoes_placas.pack(fill="x")
        acoes_intervalo = ttk.Frame(area_placas, style="Section.TFrame")
        acoes_intervalo.pack(fill="x", pady=(0, 8))

        lista_frame = ttk.Frame(area_placas, style="Section.TFrame")
        lista_frame.pack(fill="both", expand=True)
        listbox = tk.Listbox(lista_frame, selectmode=tk.EXTENDED, activestyle="none")
        scroll_lista = ttk.Scrollbar(lista_frame, orient="vertical", command=listbox.yview)
        listbox.configure(yscrollcommand=scroll_lista.set)
        listbox.pack(side="left", fill="both", expand=True)
        scroll_lista.pack(side="right", fill="y")

        ttk.Label(area_placas, text="Preview da(s) placa(s) selecionada(s):", style="SectionTitle.TLabel").pack(
            anchor="w", pady=(8, 0)
        )
        var_resumo_selecao = tk.StringVar(value="")
        ttk.Label(area_placas, textvariable=var_resumo_selecao, style="Status.TLabel").pack(anchor="w", pady=(2, 0))
        preview = tk.Text(area_placas, height=12, wrap="word", state="disabled")
        preview.pack(fill="both", expand=False, pady=(4, 0))

        def set_preview_text(texto: str) -> None:
            preview.configure(state="normal")
            preview.delete("1.0", tk.END)
            preview.insert("1.0", texto)
            preview.configure(state="disabled")

        def texto_detalhado_placa(numero: int) -> str:
            p = placas_mapa.get(numero)
            if p is None:
                return f"Placa {numero:03d}"
            lote = p.get("lote", [])
            linhas = [f"Placa {numero:03d}"]
            if isinstance(lote, list):
                linhas.extend(resumo_lote(lote))
            return "\n".join(linhas)

        def atualizar_preview(_event=None) -> None:
            cur = [int(i) for i in listbox.curselection()]
            if len(cur) == 1 and cur[0] < len(exibidas):
                numero = exibidas[cur[0]]
                set_preview_text(texto_detalhado_placa(numero))
                return
            if len(cur) > 1:
                nums = [exibidas[i] for i in cur[:6] if i < len(exibidas)]
                linhas = [f"{len(cur)} placas selecionadas (visiveis)."]
                for n in nums:
                    linhas.append(f"- Placa {n:03d}")
                if len(cur) > 6:
                    linhas.append("...")
                set_preview_text("\n".join(linhas))
                return
            if selecionadas:
                set_preview_text(texto_detalhado_placa(min(selecionadas)))
                return
            set_preview_text("Nenhuma placa selecionada.")

        def aplicar_estado_marcacao_visivel() -> None:
            selecionadas_visiveis = [i for i, n in enumerate(exibidas) if n in selecionadas]
            listbox.selection_clear(0, tk.END)
            for idx in selecionadas_visiveis:
                listbox.selection_set(idx)

        def atualizar_contagem() -> None:
            var_contagem.set(
                f"Exibidas: {len(exibidas)} | Selecionadas: {len(selecionadas)} de {len(placas)}"
            )
            var_resumo_selecao.set(f"Faixas selecionadas: {resumir_intervalos_numericos(selecionadas)}")

        def reconstruir_lista(*_args) -> None:
            termo = normalizar_texto(var_filtro.get()).lower()
            status_atual = var_status_filtro.get()
            listbox.delete(0, tk.END)
            exibidas.clear()

            for p in placas:
                numero = int(p.get("numero", 0))
                lote = p.get("lote", [])
                resumo = str(p.get("resumo", ""))
                tem_conteudo = int(p.get("qtd_validos", 0)) > 0
                texto_busca = resumo.lower()
                if isinstance(lote, list):
                    texto_busca = " ".join(
                        [
                            resumo.lower(),
                            " ".join(normalizar_texto(i.get("descricao", "")).lower() for i in lote),
                            " ".join(normalizar_preco_str(str(i.get("preco", ""))).lower() for i in lote),
                        ]
                    )
                if termo and termo not in texto_busca:
                    continue
                if status_atual == "Com conteudo" and not tem_conteudo:
                    continue
                if status_atual == "Vazias" and tem_conteudo:
                    continue
                if status_atual == "Selecionadas" and numero not in selecionadas:
                    continue
                if status_atual == "Nao selecionadas" and numero in selecionadas:
                    continue
                item = f"Placa {numero:03d} | Itens validos: {int(p.get('qtd_validos', 0))} | {resumo}"
                listbox.insert(tk.END, item)
                exibidas.append(numero)

            aplicar_estado_marcacao_visivel()
            atualizar_contagem()
            atualizar_dashboard()
            atualizar_preview()

        def sincronizar_selecao_visivel(_event=None) -> None:
            visiveis = set(exibidas)
            selecionadas.difference_update(visiveis)
            indices = set(int(i) for i in listbox.curselection())
            for i, numero in enumerate(exibidas):
                if i in indices:
                    selecionadas.add(numero)
            atualizar_contagem()
            atualizar_dashboard()
            atualizar_preview()

        def atualizar_placas() -> None:
            nonlocal placas, placas_mapa
            produtos_atuais = obter_produtos_editados(validar=False) or []
            placas = montar_placas_para_interface(produtos_atuais)
            placas_mapa = {int(p.get("numero", 0)): p for p in placas}
            numeros_validos = {int(p.get("numero", 0)) for p in placas}
            if not selecionadas:
                selecionadas.update(numeros_validos)
            else:
                selecionadas.intersection_update(numeros_validos)
                if not selecionadas:
                    selecionadas.update(numeros_validos)
            reconstruir_lista()

        recalcular_placas_ref[0] = atualizar_placas

        def marcar_todas() -> None:
            selecionadas.clear()
            for p in placas:
                selecionadas.add(int(p.get("numero", 0)))
            reconstruir_lista()

        def desmarcar_todas() -> None:
            selecionadas.clear()
            reconstruir_lista()

        def marcar_visiveis() -> None:
            for n in exibidas:
                selecionadas.add(int(n))
            reconstruir_lista()

        def desmarcar_visiveis() -> None:
            for n in exibidas:
                selecionadas.discard(int(n))
            reconstruir_lista()

        def marcar_apenas_nao_vazias() -> None:
            selecionadas.clear()
            for p in placas:
                if int(p.get("qtd_validos", 0)) > 0:
                    selecionadas.add(int(p.get("numero", 0)))
            reconstruir_lista()

        def marcar_apenas_vazias() -> None:
            selecionadas.clear()
            for p in placas:
                if int(p.get("qtd_validos", 0)) <= 0:
                    selecionadas.add(int(p.get("numero", 0)))
            reconstruir_lista()

        def inverter_selecao() -> None:
            numeros = {int(p.get("numero", 0)) for p in placas}
            atuais = set(selecionadas)
            selecionadas.clear()
            selecionadas.update(n for n in numeros if n not in atuais)
            reconstruir_lista()

        def aplicar_faixa(marcar: bool) -> None:
            try:
                numeros = parsear_intervalos_placas(var_intervalo.get(), len(placas))
            except ValueError as exc:
                messagebox.showerror("Faixa invalida", str(exc))
                return
            if not numeros:
                messagebox.showerror("Faixa obrigatoria", "Informe placas como 1-4, 7, 10-12.")
                return
            for numero in numeros:
                if marcar:
                    selecionadas.add(numero)
                else:
                    selecionadas.discard(numero)
            reconstruir_lista()

        def limpar_filtros() -> None:
            var_filtro.set("")
            var_status_filtro.set("Todas")
            var_intervalo.set("")
            reconstruir_lista()

        ttk.Button(acoes, text="Atualizar Placas", command=atualizar_placas).pack(side="left", padx=(18, 0))
        ttk.Button(acoes_placas, text="Marcar Todas", command=marcar_todas).pack(side="left")
        ttk.Button(acoes_placas, text="Desmarcar Todas", command=desmarcar_todas).pack(side="left", padx=(8, 0))
        ttk.Button(acoes_placas, text="Marcar Exibidas", command=marcar_visiveis).pack(side="left", padx=(16, 0))
        ttk.Button(acoes_placas, text="Desmarcar Exibidas", command=desmarcar_visiveis).pack(side="left", padx=(8, 0))
        ttk.Button(acoes_placas, text="Somente Nao Vazias", command=marcar_apenas_nao_vazias).pack(
            side="left", padx=(16, 0)
        )
        ttk.Button(acoes_placas, text="Somente Vazias", command=marcar_apenas_vazias).pack(side="left", padx=(8, 0))
        ttk.Button(acoes_placas, text="Inverter Selecao", command=inverter_selecao).pack(side="left", padx=(16, 0))

        ttk.Label(acoes_intervalo, text="Faixa:").pack(side="left")
        var_intervalo = tk.StringVar(value="")
        entrada_intervalo = ttk.Entry(acoes_intervalo, textvariable=var_intervalo, width=18)
        entrada_intervalo.pack(side="left", padx=(8, 8))
        ttk.Label(acoes_intervalo, text="Ex.: 1-4, 7, 10-12", style="Status.TLabel").pack(side="left")
        ttk.Button(acoes_intervalo, text="Marcar Faixa", command=lambda: aplicar_faixa(True)).pack(
            side="left", padx=(16, 0)
        )
        ttk.Button(acoes_intervalo, text="Desmarcar Faixa", command=lambda: aplicar_faixa(False)).pack(
            side="left", padx=(8, 0)
        )
        ttk.Button(acoes_intervalo, text="Limpar Filtros", command=limpar_filtros).pack(side="left", padx=(16, 0))

        listbox.bind("<<ListboxSelect>>", sincronizar_selecao_visivel)
        var_filtro.trace_add("write", reconstruir_lista)
        var_status_filtro.trace_add("write", reconstruir_lista)
        atualizar_placas()
        entrada_filtro.focus_set()
    else:
        atualizar_dashboard()

    rodape = ttk.Frame(root, style="Page.TFrame", padding=(16, 12))
    rodape.pack(fill="x")
    var_desligar_ao_final = tk.BooleanVar(value=bool(desligar_ao_final_inicial))
    ttk.Checkbutton(
        rodape,
        text="Desligar o PC ao final da impressao",
        variable=var_desligar_ao_final,
    ).pack(side="left")

    def confirmar() -> None:
        novos = obter_produtos_editados(validar=True)
        if novos is None:
            return

        selecionadas_lista: Optional[List[int]] = None
        if permitir_selecao_placas:
            if area_placas is not None:
                # Sincroniza as placas com eventuais ajustes finais no grid.
                lotes_atuais = montar_lotes(novos, tamanho_lote=4)
                numeros_validos = {i for i in range(1, len(lotes_atuais) + 1)}
                selecionadas.intersection_update(numeros_validos)
                if not selecionadas:
                    messagebox.showerror("Selecao obrigatoria", "Selecione ao menos 1 placa para produzir.")
                    return
            selecionadas_lista = sorted(selecionadas)

        resultado["produtos"] = novos
        resultado["selecionadas"] = selecionadas_lista
        resultado["desligar_ao_final"] = bool(var_desligar_ao_final.get())
        resultado["ok"] = True
        root.destroy()

    def cancelar() -> None:
        if messagebox.askyesno("Cancelar", "Cancelar a producao?"):
            resultado["ok"] = False
            root.destroy()

    ttk.Button(rodape, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(rodape, text="Iniciar Producao", command=confirmar).pack(side="right", padx=(0, 8))

    root.bind("<Control-Return>", lambda _event: confirmar())
    root.bind("<Escape>", lambda _event: cancelar())
    root.protocol("WM_DELETE_WINDOW", cancelar)
    root.mainloop()

    if not bool(resultado.get("ok")):
        return None

    produtos_saida = list(resultado.get("produtos", []))
    selecionadas_saida = resultado.get("selecionadas")
    desligar_saida = bool(resultado.get("desligar_ao_final", False))
    if isinstance(selecionadas_saida, list):
        try:
            return produtos_saida, [int(i) for i in selecionadas_saida], desligar_saida
        except Exception:
            return produtos_saida, None, desligar_saida
    return produtos_saida, None, desligar_saida


def validar_payload_revisao_web(
    payload: Dict[str, Any],
    permitir_selecao_placas: bool,
) -> Tuple[List[Dict[str, str]], Optional[List[int]], bool]:
    produtos_brutos = payload.get("products")
    if not isinstance(produtos_brutos, list):
        raise ValueError("A interface web enviou produtos em formato invalido.")

    produtos_normalizados: List[Dict[str, str]] = []
    for idx, item in enumerate(produtos_brutos, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"O item #{idx} veio em formato invalido.")
        descricao = normalizar_texto(item.get("descricao", ""))
        preco = normalizar_preco_str(str(item.get("preco", "")))
        unidade = normalizar_unidade_saida(str(item.get("unidade", "Unid.")))
        validade_oferta = normalizar_data_oferta(item.get("validade_oferta", ""))
        codigo_barras_bruto = extrair_digitos_codigo_barras(item.get("codigo_barras", ""))
        codigo_barras = codigo_barras_bruto if len(codigo_barras_bruto) == 13 else ""
        usar_codigo_barras = normalizar_flag_codigo_barras_saida(
            item.get("usar_codigo_barras", False),
            codigo_barras,
        )

        if not descricao and not preco:
            continue
        if not descricao:
            raise ValueError(f"O produto #{idx} esta sem descricao.")
        if not preco:
            raise ValueError(f"O produto #{idx} esta sem preco valido.")
        if codigo_barras_bruto and len(codigo_barras_bruto) != 13:
            raise ValueError(f"O produto #{idx} esta com codigo de barras invalido. Use 13 digitos.")
        produtos_normalizados.append(
            {
                "descricao": descricao,
                "preco": preco,
                "unidade": unidade,
                "validade_oferta": validade_oferta,
                "codigo_barras": codigo_barras,
                "usar_codigo_barras": bool(usar_codigo_barras),
            }
        )

    if not produtos_normalizados:
        raise ValueError("Nenhum produto valido foi informado.")

    selecionadas_saida: Optional[List[int]] = None
    if permitir_selecao_placas:
        bruto_selecionadas = payload.get("selected_plates")
        if not isinstance(bruto_selecionadas, list):
            raise ValueError("A interface web nao informou as placas selecionadas.")
        lotes_atuais = montar_lotes(produtos_normalizados, tamanho_lote=4)
        max_placas = len(lotes_atuais)
        selecionadas_limpas = sorted(
            {
                int(numero)
                for numero in bruto_selecionadas
                if str(numero).strip().isdigit() and int(numero) > 0
            }
        )
        if not selecionadas_limpas:
            raise ValueError("Selecione ao menos 1 placa para produzir.")
        if any(numero > max_placas for numero in selecionadas_limpas):
            raise ValueError("A selecao de placas ficou fora do intervalo atual.")
        selecionadas_saida = selecionadas_limpas

    desligar_saida = bool(payload.get("shutdown_after_print", False))
    return produtos_normalizados, selecionadas_saida, desligar_saida


def revisar_produtos_navegador(
    produtos: List[Dict[str, str]],
    permitir_selecao_placas: bool = True,
    desligar_ao_final_inicial: bool = False,
    permitir_analise_entrada: bool = False,
    analisar_entrada_callback: Optional[Callable[[Dict[str, Any]], Dict[str, Any]]] = None,
    estado_revisao: Optional[Dict[str, Any]] = None,
    produtos_originais_iniciais: Optional[List[Dict[str, str]]] = None,
    usuario_logado: Optional[Dict[str, Any]] = None,
    usuarios_acesso: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Tuple[List[Dict[str, str]], Optional[List[int]], bool]]:
    pagina_html = carregar_pagina_configuracao_web()
    usuarios_acesso_validos = [
        dict(item) for item in usuarios_acesso if isinstance(item, dict)
    ] if isinstance(usuarios_acesso, list) else []
    usuario_logado_atual = dict(usuario_logado) if isinstance(usuario_logado, dict) else None

    def obter_usuario_logado_serializado() -> Dict[str, str]:
        return serializar_usuario_logado(usuario_logado_atual)

    def obter_usuario_ator() -> Dict[str, Any]:
        return dict(usuario_logado_atual) if isinstance(usuario_logado_atual, dict) else {}

    produtos_iniciais = [
        {
            "descricao": normalizar_texto(p.get("descricao", "")),
            "preco": normalizar_preco_str(str(p.get("preco", ""))),
            "unidade": normalizar_unidade_saida(str(p.get("unidade", "Unid."))),
            "validade_oferta": normalizar_data_oferta(p.get("validade_oferta", "")),
            "codigo_barras": normalizar_codigo_barras_saida(p.get("codigo_barras", "")),
            "usar_codigo_barras": normalizar_flag_codigo_barras_saida(
                p.get("usar_codigo_barras", False),
                p.get("codigo_barras", ""),
            ),
        }
        for p in produtos
    ]
    if isinstance(produtos_originais_iniciais, list) and produtos_originais_iniciais:
        produtos_brutos_atual = [dict(p) for p in produtos_originais_iniciais if isinstance(p, dict)]
    else:
        produtos_brutos_atual = [dict(p) for p in produtos_iniciais]
    origem_atual = "produtos atuais"
    stats_atuais: Dict[str, int] = {}
    estado_inicial = {
        "products": produtos_iniciais,
        "allow_plate_selection": bool(permitir_selecao_placas),
        "shutdown_after_print": bool(desligar_ao_final_inicial),
        "allow_input_analysis": bool(permitir_analise_entrada and callable(analisar_entrada_callback)),
        "analysis_origin": origem_atual,
        "analysis_stats": stats_atuais,
        "initial_plate_summary": montar_placas_para_interface(produtos_iniciais),
        "logged_user": obter_usuario_logado_serializado(),
    }
    estado_json = json.dumps(estado_inicial, ensure_ascii=False).replace("</", "<\\/")

    resultado: Dict[str, Any] = {"status": "pending", "payload": None}
    finalizado = threading.Event()
    progress_state: Dict[str, Any] = {
        "phase": "editing",
        "message": "Revise os itens e clique em Iniciar producao.",
        "started_at": None,
        "finished_at": None,
        "updated_at": time.time(),
        "total_plates": 0,
        "completed_plates": 0,
        "current_plate": 0,
        "current_plate_label": "",
        "current_product": "",
        "eta_seconds": None,
        "elapsed_seconds": 0.0,
    }
    class RevisaoHandler(BaseHTTPRequestHandler):
        def log_message(self, _format: str, *_args) -> None:
            return

        def _write_json(self, status: int, payload: Dict[str, Any]) -> None:
            corpo = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(corpo)))
            self.end_headers()
            self.wfile.write(corpo)

        def do_GET(self) -> None:
            if self.path == "/session-user":
                self._write_json(HTTPStatus.OK, {"ok": True, "user": obter_usuario_logado_serializado()})
                return
            if self.path == "/progress":
                now_ts = time.time()
                started_at = progress_state.get("started_at")
                if isinstance(started_at, (int, float)) and float(started_at) > 0:
                    progress_state["elapsed_seconds"] = max(0.0, now_ts - float(started_at))
                progress_state["updated_at"] = now_ts
                self._write_json(HTTPStatus.OK, {"ok": True, "progress": progress_state})
                return
            if self.path not in ("/", "/index.html"):
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            corpo = pagina_html.replace("__INITIAL_STATE__", estado_json).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(corpo)))
            self.end_headers()
            self.wfile.write(corpo)

        def do_POST(self) -> None:
            nonlocal usuario_logado_atual
            nonlocal origem_atual, stats_atuais
            try:
                content_length = int(self.headers.get("Content-Length", "0") or 0)
            except Exception:
                content_length = 0
            bruto = self.rfile.read(max(0, content_length))
            try:
                payload = json.loads(bruto.decode("utf-8"))
            except Exception:
                self._write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "JSON invalido."})
                return

            if self.path == "/submit":
                usuario_ator = obter_usuario_ator()
                if not usuario_pode_gerar_placas(usuario_ator):
                    self._write_json(
                        HTTPStatus.FORBIDDEN,
                        {"ok": False, "error": "Entre com um usuario valido antes de iniciar a producao."},
                    )
                    return
                try:
                    produtos_saida, selecionadas_saida, desligar_saida = validar_payload_revisao_web(
                        payload if isinstance(payload, dict) else {},
                        permitir_selecao_placas=permitir_selecao_placas,
                    )
                except Exception as exc:
                    self._write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
                    return
                resultado["status"] = "ok"
                resultado["payload"] = (produtos_saida, selecionadas_saida, desligar_saida)
                total_plates = len(montar_lotes(produtos_saida, tamanho_lote=4))
                if isinstance(selecionadas_saida, list) and selecionadas_saida:
                    total_plates = len(selecionadas_saida)
                progress_state.update(
                    {
                        "phase": "queued",
                        "message": "Configuracao recebida. Aguardando inicio da producao...",
                        "started_at": None,
                        "finished_at": None,
                        "updated_at": time.time(),
                        "total_plates": int(total_plates),
                        "completed_plates": 0,
                        "current_plate": 0,
                        "current_plate_label": "",
                        "current_product": "",
                        "eta_seconds": None,
                        "elapsed_seconds": 0.0,
                    }
                )
                if estado_revisao is not None:
                    estado_revisao["produtos_originais_extraidos"] = [dict(item) for item in produtos_brutos_atual]
                    estado_revisao["origem_entrada"] = origem_atual
                    estado_revisao["web_progress_state"] = progress_state
                    estado_revisao["usuario_logado"] = dict(usuario_ator)
                finalizado.set()
                self._write_json(HTTPStatus.OK, {"ok": True, "message": "Configuracao recebida."})
                return

            if self.path == "/analyze":
                if not (permitir_analise_entrada and callable(analisar_entrada_callback)):
                    self._write_json(HTTPStatus.NOT_FOUND, {"ok": False, "error": "Analise indisponivel."})
                    return
                try:
                    analysis_result = analisar_entrada_callback(payload if isinstance(payload, dict) else {})
                except EntradaInvalida as exc:
                    self._write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
                    return
                except Exception as exc:
                    self._write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
                    return

                produtos_processados = analysis_result.get("products", [])
                produtos_brutos = analysis_result.get("raw_products", [])
                if not isinstance(produtos_processados, list) or not isinstance(produtos_brutos, list):
                    self._write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": "Resposta de analise invalida."})
                    return

                produtos_brutos_atual[:] = [dict(item) for item in produtos_brutos if isinstance(item, dict)]
                origem_atual_local = normalizar_texto(analysis_result.get("origin", "")) or "pagina web"
                stats_local = analysis_result.get("stats", {})
                try:
                    plate_summary = analysis_result.get("plate_summary") or montar_placas_para_interface(produtos_processados)
                except Exception:
                    plate_summary = montar_placas_para_interface(produtos_processados)

                origem_atual = origem_atual_local
                stats_atuais = dict(stats_local) if isinstance(stats_local, dict) else {}
                self._write_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "origin": origem_atual,
                        "stats": stats_atuais,
                        "products": produtos_processados,
                        "plate_summary": plate_summary,
                    },
                )
                return

            if self.path == "/login":
                if not usuarios_acesso_validos:
                    self._write_json(
                        HTTPStatus.SERVICE_UNAVAILABLE,
                        {"ok": False, "error": "Troca de usuario indisponivel nesta execucao."},
                    )
                    return
                dados = payload if isinstance(payload, dict) else {}
                usuario = normalizar_texto(str(dados.get("usuario", "")))
                senha = str(dados.get("senha", "")).strip()
                lembrar_24h = bool(dados.get("remember_24h", False))
                user_auth = autenticar_login(usuarios_acesso_validos, usuario, senha)
                if user_auth is None:
                    self._write_json(
                        HTTPStatus.UNAUTHORIZED,
                        {"ok": False, "error": "Usuario, e-mail, telefone ou senha invalidos."},
                    )
                    return
                if not usuario_pode_gerar_placas(user_auth):
                    self._write_json(
                        HTTPStatus.FORBIDDEN,
                        {
                            "ok": False,
                            "error": "Seu nivel nao possui permissao para geracao de placas.",
                        },
                    )
                    return
                if lembrar_24h:
                    try:
                        salvar_sessao_login_24h(user_auth)
                    except Exception as exc:
                        self._write_json(
                            HTTPStatus.INTERNAL_SERVER_ERROR,
                            {"ok": False, "error": f"Nao foi possivel salvar a sessao de 24 horas ({exc})."},
                        )
                        return
                else:
                    limpar_sessao_login_24h()
                usuario_logado_atual = dict(user_auth)
                self._write_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "message": f"Acesso alterado para {normalizar_texto(str(user_auth.get('nome') or user_auth.get('usuario') or ''))}.",
                        "user": obter_usuario_logado_serializado(),
                    },
                )
                return

            if self.path == "/users/create":
                usuario_ator = obter_usuario_ator()
                if not usuario_pode_gerenciar_usuarios(usuario_ator):
                    self._write_json(
                        HTTPStatus.FORBIDDEN,
                        {"ok": False, "error": "Apenas usuarios Desenvolvedor podem cadastrar novos usuarios."},
                    )
                    return
                token_github = normalizar_texto(os.environ.get(GITHUB_TOKEN_ENV, ""))
                if not token_github:
                    self._write_json(
                        HTTPStatus.FORBIDDEN,
                        {
                            "ok": False,
                            "error": (
                                "Cadastro indisponivel: defina a variavel "
                                f"{GITHUB_TOKEN_ENV} com permissao de escrita no repositorio."
                            ),
                        },
                    )
                    return
                dados = payload if isinstance(payload, dict) else {}
                try:
                    usuario_criado = criar_usuario_acesso_github(
                        token=token_github,
                        repo=DEFAULT_GITHUB_LOG_REPO,
                        caminho_arquivo=DEFAULT_GITHUB_USERS_PATH,
                        branch=DEFAULT_GITHUB_LOG_BRANCH,
                        criado_por=str(usuario_ator.get("usuario") or usuario_ator.get("nome") or ""),
                        novo_usuario=dados,
                    )
                except ValueError as exc:
                    self._write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": str(exc)})
                    return
                except Exception as exc:
                    self._write_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})
                    return

                self._write_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "message": "Usuario cadastrado com sucesso no GitHub.",
                        "user": usuario_criado,
                    },
                )
                return

            if self.path == "/cancel":
                resultado["status"] = "cancelled"
                finalizado.set()
                self._write_json(HTTPStatus.OK, {"ok": True})
                return

            if self.path == "/logout":
                limpar_sessao_login_24h()
                usuario_logado_atual = None
                self._write_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "message": "Sessao encerrada. Entre com outro usuario para continuar.",
                        "user": obter_usuario_logado_serializado(),
                    },
                )
                return

            self.send_error(HTTPStatus.NOT_FOUND)

    servidor = ThreadingHTTPServer(("127.0.0.1", 0), RevisaoHandler)
    thread_servidor = threading.Thread(target=servidor.serve_forever, daemon=True)
    thread_servidor.start()

    url = f"http://127.0.0.1:{int(servidor.server_port)}/"
    print(f"[WEB] Configuracao: abra {url}")
    try:
        webbrowser.open(url, new=2, autoraise=True)
    except Exception:
        pass

    try:
        finalizado.wait()
    finally:
        if estado_revisao is None:
            try:
                servidor.shutdown()
            except Exception:
                pass
            try:
                servidor.server_close()
            except Exception:
                pass
            thread_servidor.join(timeout=2.0)
        else:
            estado_revisao["web_server"] = servidor
            estado_revisao["web_server_thread"] = thread_servidor

    if resultado.get("status") != "ok":
        return None
    payload_final = resultado.get("payload")
    if not isinstance(payload_final, tuple) or len(payload_final) != 3:
        return None
    return payload_final  # type: ignore[return-value]


def autenticar_acesso_navegador(
    usuarios: List[Dict[str, Any]],
    source_label: str,
) -> Optional[Dict[str, Any]]:
    pagina_html = carregar_pagina_login_web()
    suporte = None
    for usuario in usuarios:
        if normalizar_usuario_login(str(usuario.get("usuario", ""))) == normalizar_usuario_login("Jeferson"):
            suporte = usuario
            break
    if suporte is None and usuarios:
        suporte = usuarios[0]
    if suporte is None:
        suporte = {"nome": "Suporte", "perfil": "Administrador", "telefone": "", "email": ""}

    estado_inicial = {
        "source_label": source_label,
        "users_count": len(usuarios),
        "developer_name": normalizar_texto(str(suporte.get("nome") or suporte.get("usuario") or "Suporte")),
        "developer_role": normalizar_texto(str(suporte.get("perfil") or "Administrador")),
        "developer_phone": normalizar_texto(str(suporte.get("telefone") or "")),
        "developer_email": normalizar_texto(str(suporte.get("email") or "")),
        "session_hours": LOGIN_SESSION_HOURS,
    }
    estado_json = json.dumps(estado_inicial, ensure_ascii=False).replace("</", "<\\/")
    resultado: Dict[str, Any] = {"status": "pending", "user": None}
    finalizado = threading.Event()

    class LoginHandler(BaseHTTPRequestHandler):
        def log_message(self, _format: str, *_args) -> None:
            return

        def _write_json(self, status: int, payload: Dict[str, Any]) -> None:
            corpo = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(corpo)))
            self.end_headers()
            self.wfile.write(corpo)

        def do_GET(self) -> None:
            if self.path not in ("/", "/index.html"):
                self.send_error(HTTPStatus.NOT_FOUND)
                return
            corpo = pagina_html.replace("__INITIAL_STATE__", estado_json).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(corpo)))
            self.end_headers()
            self.wfile.write(corpo)

        def do_POST(self) -> None:
            try:
                content_length = int(self.headers.get("Content-Length", "0") or 0)
            except Exception:
                content_length = 0
            bruto = self.rfile.read(max(0, content_length))
            try:
                payload = json.loads(bruto.decode("utf-8"))
            except Exception:
                self._write_json(HTTPStatus.BAD_REQUEST, {"ok": False, "error": "JSON invalido."})
                return

            if self.path == "/login":
                dados = payload if isinstance(payload, dict) else {}
                usuario = normalizar_texto(str(dados.get("usuario", "")))
                senha = str(dados.get("senha", "")).strip()
                lembrar_24h = bool(dados.get("remember_24h", False))
                user_auth = autenticar_login(usuarios, usuario, senha)
                if user_auth is None:
                    self._write_json(
                        HTTPStatus.UNAUTHORIZED,
                        {"ok": False, "error": "Usuario, e-mail, telefone ou senha invalidos."},
                    )
                    return
                if not usuario_pode_gerar_placas(user_auth):
                    self._write_json(
                        HTTPStatus.FORBIDDEN,
                        {
                            "ok": False,
                            "error": "Seu nivel nao possui permissao para geracao de placas.",
                        },
                    )
                    return
                if lembrar_24h:
                    try:
                        salvar_sessao_login_24h(user_auth)
                    except Exception as exc:
                        self._write_json(
                            HTTPStatus.INTERNAL_SERVER_ERROR,
                            {"ok": False, "error": f"Nao foi possivel salvar a sessao de 24 horas ({exc})."},
                        )
                        return
                else:
                    limpar_sessao_login_24h()
                resultado["status"] = "ok"
                resultado["user"] = user_auth
                finalizado.set()
                self._write_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "nome": normalizar_texto(str(user_auth.get("nome") or user_auth.get("usuario") or "")),
                        "perfil": normalizar_texto(str(user_auth.get("perfil") or "")),
                        "remember_24h": lembrar_24h,
                    },
                )
                return

            if self.path == "/cancel":
                resultado["status"] = "cancelled"
                finalizado.set()
                self._write_json(HTTPStatus.OK, {"ok": True})
                return

            self.send_error(HTTPStatus.NOT_FOUND)

    servidor = ThreadingHTTPServer(("127.0.0.1", 0), LoginHandler)
    thread_servidor = threading.Thread(target=servidor.serve_forever, daemon=True)
    thread_servidor.start()

    url = f"http://127.0.0.1:{int(servidor.server_port)}/"
    print(f"[WEB] Login: abra {url}")
    try:
        webbrowser.open(url, new=2, autoraise=True)
    except Exception:
        pass

    try:
        finalizado.wait()
    finally:
        try:
            servidor.shutdown()
        except Exception:
            pass
        try:
            servidor.server_close()
        except Exception:
            pass
        thread_servidor.join(timeout=2.0)

    if resultado.get("status") != "ok":
        return None
    usuario_logado = resultado.get("user")
    if not isinstance(usuario_logado, dict):
        return None
    return usuario_logado


def revisar_produtos_interface(
    produtos: List[Dict[str, str]],
    permitir_selecao_placas: bool = True,
    desligar_ao_final_inicial: bool = False,
    permitir_analise_entrada: bool = False,
    analisar_entrada_callback: Optional[Callable[[Dict[str, Any]], Dict[str, Any]]] = None,
    estado_revisao: Optional[Dict[str, Any]] = None,
    produtos_originais_iniciais: Optional[List[Dict[str, str]]] = None,
    usuario_logado: Optional[Dict[str, Any]] = None,
    usuarios_acesso: Optional[List[Dict[str, Any]]] = None,
) -> Optional[Tuple[List[Dict[str, str]], Optional[List[int]], bool]]:
    try:
        return revisar_produtos_navegador(
            produtos,
            permitir_selecao_placas=permitir_selecao_placas,
            desligar_ao_final_inicial=desligar_ao_final_inicial,
            permitir_analise_entrada=permitir_analise_entrada,
            analisar_entrada_callback=analisar_entrada_callback,
            estado_revisao=estado_revisao,
            produtos_originais_iniciais=produtos_originais_iniciais,
            usuario_logado=usuario_logado,
            usuarios_acesso=usuarios_acesso,
        )
    except Exception as exc:
        print(f"Aviso: falha ao abrir a configuracao web ({exc}). Voltando para a interface local.")
        return revisar_produtos_tela(
            produtos,
            permitir_selecao_placas=permitir_selecao_placas,
            desligar_ao_final_inicial=desligar_ao_final_inicial,
        )


def parser_args() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Atualiza Validade.cdr a partir de Excel/PDF em lotes de 4 produtos"
    )
    p.add_argument("--entrada", required=False, help="Arquivo de entrada (.xlsx/.xlsm/.pdf/.txt/.md)")
    p.add_argument("--texto-bruto", default=None, help="Texto bruto com descricao e preco para importar direto.")
    p.add_argument("--arquivo-cdr", required=True, help="Template CDR")
    p.add_argument("--salvar-em", default=None, help="Destino do CDR de trabalho")
    p.add_argument("--sem-copia", action="store_true", help="Usa o proprio CDR sem criar copia")
    p.add_argument(
        "--salvar-cdr",
        action="store_true",
        help="Salva alteracoes no CDR. Padrao: nao salvar (somente imprimir).",
    )
    p.add_argument("--nao-imprimir", action="store_true", help="Atualiza lotes sem mandar para impressora")
    p.add_argument(
        "--sem-confirmacao-impressao",
        action="store_true",
        help="Nao pede confirmacao apos cada impressao de placa.",
    )
    p.add_argument("--copias", type=int, default=1, help="Copias por lote")
    p.add_argument("--impressora", default=None, help="Nome da impressora (opcional)")
    p.add_argument("--pausa-segundos", type=float, default=0.0, help="Pausa entre lotes")
    p.add_argument(
        "--desligar-ao-final",
        action="store_true",
        help="Agenda o desligamento do Windows ao final da impressao de todas as placas.",
    )
    p.add_argument(
        "--arquivo-aprendizado",
        default=DB_APRENDIZADO_ARQUIVO,
        help="Arquivo JSON para memoria de correcoes (descricao/unidade).",
    )
    p.add_argument(
        "--arquivo-velocidade",
        default=DB_VELOCIDADE_ARQUIVO,
        help="Arquivo JSON para memoria de performance (modo rapido inteligente).",
    )
    p.add_argument(
        "--arquivo-corretor-acentos",
        default=DB_CORRETOR_ACENTOS_ARQUIVO,
        help="Arquivo JSON com dicionario de palavras para corrigir acentos.",
    )
    p.add_argument(
        "--sem-aprendizado",
        action="store_true",
        help="Desabilita leitura e gravacao do aprendizado (correcoes e velocidade).",
    )
    p.add_argument(
        "--sem-corretor-acentos",
        action="store_true",
        help="Desabilita a correcao automatica de acentos nas descricoes.",
    )
    p.add_argument(
        "--usar-ia-local",
        action="store_true",
        help="Ativa limpeza inteligente opcional com IA local via Ollama antes da revisao.",
    )
    p.add_argument(
        "--modelo-ia-local",
        default=DEFAULT_OLLAMA_MODEL,
        help=f"Modelo do Ollama para limpeza local. Padrao: {DEFAULT_OLLAMA_MODEL}.",
    )
    p.add_argument(
        "--timeout-ia-local",
        type=float,
        default=DEFAULT_OLLAMA_TIMEOUT_SECONDS,
        help=f"Tempo maximo por item na IA local. Padrao: {DEFAULT_OLLAMA_TIMEOUT_SECONDS}s.",
    )
    p.add_argument(
        "--max-itens-ia-local",
        type=int,
        default=DEFAULT_OLLAMA_MAX_ITEMS,
        help=f"Quantidade maxima de itens enviados para IA local por execucao. Padrao: {DEFAULT_OLLAMA_MAX_ITEMS}.",
    )
    p.add_argument(
        "--sem-modo-rapido-inteligente",
        action="store_true",
        help="Desabilita reutilizacao inteligente da sessao Corel entre placas.",
    )
    p.add_argument(
        "--sem-tela-unidade",
        action="store_true",
        help="Nao abre a tela de revisao (descricao + Kg/Unid por produto).",
    )
    p.add_argument(
        "--sem-tela-placas",
        action="store_true",
        help="Nao abre a tela para selecionar quais placas (lotes) produzir.",
    )
    p.add_argument(
        "--sem-log-github",
        action="store_true",
        help="Opcao legada (ignorada). O log remoto das placas no GitHub e obrigatorio.",
    )
    return p


def resumo_lote(lote: List[Dict[str, str]]) -> List[str]:
    linhas: List[str] = []
    for i, p in enumerate(lote, start=1):
        desc = p.get("descricao", "").strip()
        preco = p.get("preco", "").strip()
        unid = p.get("unidade", "").strip()
        validade = normalizar_data_oferta(p.get("validade_oferta", ""))
        if not desc and not preco:
            linhas.append(f"  {i}. [vazio]")
            continue
        sufixo_validade = f" | Validade {validade}" if validade else ""
        linhas.append(f"  {i}. {desc} | {preco} | {unid or 'Unid.'}{sufixo_validade}")
    return linhas


def resumo_lote_curto(lote: List[Dict[str, str]], max_itens: int = 2) -> str:
    itens: List[str] = []
    for p in lote:
        desc = normalizar_texto(p.get("descricao", ""))
        preco = normalizar_preco_str(str(p.get("preco", "")))
        validade = normalizar_data_oferta(p.get("validade_oferta", ""))
        if not desc:
            continue
        sufixo = f" ate {validade}" if validade else ""
        itens.append(f"{desc} ({preco}){sufixo}" if preco else f"{desc}{sufixo}")
        if len(itens) >= max_itens:
            break
    if not itens:
        return "[vazio]"
    return " | ".join(itens)


def contar_produtos_validos(produtos: List[Dict[str, str]]) -> int:
    return sum(
        1
        for p in produtos
        if normalizar_texto(p.get("descricao", "")) and normalizar_preco_str(str(p.get("preco", "")))
    )


def montar_placas_para_interface(produtos: List[Dict[str, str]], tamanho_lote: int = 4) -> List[Dict[str, object]]:
    lotes = montar_lotes(produtos, tamanho_lote=tamanho_lote)
    placas: List[Dict[str, object]] = []
    for i, lote in enumerate(lotes, start=1):
        placas.append(
            {
                "numero": i,
                "lote": lote,
                "qtd_validos": contar_produtos_validos(lote),
                "resumo": resumo_lote_curto(lote, max_itens=3),
            }
        )
    return placas


def configurar_estilo_interface(root, ttk) -> None:
    style = ttk.Style(root)
    try:
        if "clam" in style.theme_names():
            style.theme_use("clam")
    except Exception:
        pass

    root.configure(bg="#ebe7df")
    style.configure("Page.TFrame", background="#ebe7df")
    style.configure("Hero.TFrame", background="#0d1b34")
    style.configure(
        "HeroTitle.TLabel",
        background="#0d1b34",
        foreground="#fff9f0",
        font=("Bahnschrift SemiBold", 18),
    )
    style.configure(
        "HeroText.TLabel",
        background="#0d1b34",
        foreground="#d7dfeb",
        font=("Segoe UI", 10),
    )
    style.configure(
        "HeroMeta.TLabel",
        background="#0d1b34",
        foreground="#f8c27c",
        font=("Segoe UI Semibold", 9),
    )
    style.configure("Card.TFrame", background="#fffdfa", relief="solid", borderwidth=1)
    style.configure(
        "CardTitle.TLabel",
        background="#fffdfa",
        foreground="#6d7785",
        font=("Segoe UI Semibold", 9),
    )
    style.configure(
        "CardValue.TLabel",
        background="#fffdfa",
        foreground="#11223b",
        font=("Bahnschrift SemiBold", 17),
    )
    style.configure("Section.TFrame", background="#fffdfa", relief="solid", borderwidth=1)
    style.configure(
        "SectionTitle.TLabel",
        background="#fffdfa",
        foreground="#102033",
        font=("Bahnschrift SemiBold", 12),
    )
    style.configure(
        "SectionText.TLabel",
        background="#fffdfa",
        foreground="#556274",
        font=("Segoe UI", 9),
    )
    style.configure(
        "Muted.TLabel",
        background="#ebe7df",
        foreground="#667487",
        font=("Segoe UI", 9),
    )
    style.configure(
        "Status.TLabel",
        background="#ebe7df",
        foreground="#556274",
        font=("Segoe UI Semibold", 9),
    )
    style.configure(
        "Accent.TButton",
        background="#ff6b35",
        foreground="#fffaf2",
        borderwidth=0,
        focusthickness=0,
        padding=(16, 10),
        font=("Segoe UI Semibold", 10),
    )
    style.map(
        "Accent.TButton",
        background=[("active", "#f45d26"), ("pressed", "#df4f1b")],
        foreground=[("disabled", "#f9dccf")],
    )
    style.configure(
        "Soft.TButton",
        background="#fffdfa",
        foreground="#102033",
        borderwidth=1,
        padding=(14, 10),
        font=("Segoe UI Semibold", 10),
    )
    style.map(
        "Soft.TButton",
        background=[("active", "#f4efe6"), ("pressed", "#eee6d7")],
    )
    style.configure(
        "Picker.TRadiobutton",
        background="#fffdfa",
        foreground="#102033",
        font=("Segoe UI Semibold", 10),
        padding=(12, 8),
    )
    style.map(
        "Picker.TRadiobutton",
        background=[("active", "#f5efe7"), ("selected", "#f3eadf")],
        foreground=[("selected", "#0d1b34")],
    )
    style.configure(
        "TEntry",
        fieldbackground="#ffffff",
        foreground="#102033",
        bordercolor="#d8d0c4",
        lightcolor="#d8d0c4",
        darkcolor="#d8d0c4",
        padding=9,
    )
    style.map(
        "TEntry",
        bordercolor=[("focus", "#0f766e")],
        lightcolor=[("focus", "#0f766e")],
        darkcolor=[("focus", "#0f766e")],
    )


def resumir_intervalos_numericos(numeros: Sequence[int]) -> str:
    itens = sorted({int(n) for n in numeros if int(n) > 0})
    if not itens:
        return "nenhuma"

    faixas: List[str] = []
    inicio = itens[0]
    fim = itens[0]
    for numero in itens[1:]:
        if numero == fim + 1:
            fim = numero
            continue
        faixas.append(f"{inicio}" if inicio == fim else f"{inicio}-{fim}")
        inicio = numero
        fim = numero
    faixas.append(f"{inicio}" if inicio == fim else f"{inicio}-{fim}")
    return ", ".join(faixas)


def parsear_intervalos_placas(texto: str, max_numero: int) -> List[int]:
    numeros: set[int] = set()
    partes = [p.strip() for p in normalizar_texto(texto).split(",") if p.strip()]
    if not partes:
        return []

    for parte in partes:
        if "-" in parte:
            pedacos = [p.strip() for p in parte.split("-", 1)]
            if len(pedacos) != 2 or not pedacos[0].isdigit() or not pedacos[1].isdigit():
                raise ValueError("Use faixas como 1-4, 7, 10-12.")
            inicio = int(pedacos[0])
            fim = int(pedacos[1])
            if inicio <= 0 or fim <= 0 or inicio > fim:
                raise ValueError("As faixas precisam usar numeros positivos em ordem crescente.")
            for numero in range(inicio, fim + 1):
                if numero > max_numero:
                    raise ValueError(f"A placa {numero} nao existe neste lote.")
                numeros.add(numero)
            continue

        if not parte.isdigit():
            raise ValueError("Use numeros separados por virgula ou faixas com hifen.")
        numero = int(parte)
        if numero <= 0 or numero > max_numero:
            raise ValueError(f"A placa {numero} nao existe neste lote.")
        numeros.add(numero)

    return sorted(numeros)


def agendar_desligamento_windows(segundos: int = 60) -> Tuple[bool, str]:
    atraso = max(0, int(segundos))
    if sys.platform != "win32":
        return False, "Desligamento automatico suportado apenas no Windows."

    try:
        subprocess.run(
            [
                "shutdown",
                "/s",
                "/t",
                str(atraso),
                "/c",
                "Automacao de ofertas concluida",
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        if atraso <= 0:
            return True, "Desligamento do Windows acionado imediatamente."
        return True, f"Desligamento do Windows agendado para {atraso} segundos. Use 'shutdown /a' para cancelar."
    except Exception as exc:
        return False, f"Nao foi possivel agendar o desligamento automatico ({exc})."


def selecionar_placas_tela(
    lotes: List[List[Dict[str, str]]],
    desligar_ao_final_inicial: bool = False,
) -> Optional[Tuple[List[int], bool]]:
    try:
        import tkinter as tk
        from tkinter import messagebox, ttk
    except Exception:
        print("Aviso: tkinter indisponivel. Todas as placas serao produzidas.")
        return list(range(1, len(lotes) + 1)), bool(desligar_ao_final_inicial)

    resultado: Dict[str, object] = {
        "ok": False,
        "selecionadas": [],
        "desligar_ao_final": bool(desligar_ao_final_inicial),
    }

    placas = montar_placas_para_interface([p for lote in lotes for p in lote], tamanho_lote=4)
    selecionadas = {int(p["numero"]) for p in placas}
    exibidas: List[int] = []

    root = tk.Tk()
    root.title("Selecao de Placas")
    root.geometry("1280x760")
    root.minsize(1080, 620)
    configurar_estilo_interface(root, ttk)

    cabecalho = ttk.Frame(root, style="Hero.TFrame", padding=(18, 18))
    cabecalho.pack(fill="x")
    ttk.Label(cabecalho, text="Selecao de placas para producao", style="HeroTitle.TLabel").pack(anchor="w")
    ttk.Label(
        cabecalho,
        text=(
            "Filtre os lotes, confira o preview e marque apenas as placas que devem seguir para o CorelDRAW."
        ),
        style="HeroText.TLabel",
        wraplength=1180,
        justify="left",
    ).pack(anchor="w", pady=(6, 0))

    faixa_resumo = ttk.Frame(root, style="Page.TFrame", padding=(16, 12, 16, 4))
    faixa_resumo.pack(fill="x")

    def criar_indicador(parent, titulo: str) -> object:
        card = ttk.Frame(parent, style="Card.TFrame", padding=(14, 10))
        card.pack(side="left", fill="x", expand=True, padx=(0, 10))
        ttk.Label(card, text=titulo, style="CardTitle.TLabel").pack(anchor="w")
        var = tk.StringVar(value="0")
        ttk.Label(card, textvariable=var, style="CardValue.TLabel").pack(anchor="w", pady=(4, 0))
        return var

    var_total_placas = criar_indicador(faixa_resumo, "Placas")
    var_selecionadas = criar_indicador(faixa_resumo, "Selecionadas")
    var_com_conteudo = criar_indicador(faixa_resumo, "Com conteudo")
    var_vazias = criar_indicador(faixa_resumo, "Vazias")

    filtros = ttk.Frame(root, style="Page.TFrame", padding=(16, 0, 16, 0))
    filtros.pack(fill="x")
    ttk.Label(filtros, text="Filtro:").pack(side="left")
    var_filtro = tk.StringVar(value="")
    entrada_filtro = ttk.Entry(filtros, textvariable=var_filtro, width=52)
    entrada_filtro.pack(side="left", padx=(8, 8))
    ttk.Label(filtros, text="Status:").pack(side="left")
    var_status_filtro = tk.StringVar(value="Todas")
    combo_status = ttk.Combobox(
        filtros,
        textvariable=var_status_filtro,
        values=("Todas", "Com conteudo", "Vazias", "Selecionadas", "Nao selecionadas"),
        width=17,
        state="readonly",
    )
    combo_status.pack(side="left", padx=(8, 8))
    var_contagem = tk.StringVar(value="")
    ttk.Label(filtros, textvariable=var_contagem, style="Status.TLabel").pack(side="left")

    acoes = ttk.Frame(root, style="Page.TFrame", padding=(16, 10, 16, 8))
    acoes.pack(fill="x")
    acoes_intervalo = ttk.Frame(root, style="Page.TFrame", padding=(16, 0, 16, 8))
    acoes_intervalo.pack(fill="x")

    conteudo = ttk.Frame(root, style="Page.TFrame", padding=(16, 4, 16, 0))
    conteudo.pack(fill="both", expand=True)
    esquerda = ttk.Frame(conteudo, style="Section.TFrame", padding=(14, 14))
    esquerda.pack(side="left", fill="both", expand=True)
    direita = ttk.Frame(conteudo, style="Section.TFrame", padding=(14, 14))
    direita.pack(side="left", fill="both", expand=True, padx=(12, 0))

    ttk.Label(esquerda, text="Lista de placas", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        esquerda,
        text="A lista respeita o filtro e preserva a selecao atual enquanto voce navega entre os lotes.",
        style="SectionText.TLabel",
        wraplength=520,
        justify="left",
    ).pack(anchor="w", pady=(4, 10))

    listbox = tk.Listbox(esquerda, selectmode=tk.EXTENDED, activestyle="none")
    scroll = ttk.Scrollbar(esquerda, orient="vertical", command=listbox.yview)
    listbox.configure(yscrollcommand=scroll.set)
    listbox.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    ttk.Label(direita, text="Preview das placas", style="SectionTitle.TLabel").pack(anchor="w")
    ttk.Label(
        direita,
        text="Visualize os itens do lote destacado ou um resumo rapido das placas visiveis selecionadas.",
        style="SectionText.TLabel",
        wraplength=520,
        justify="left",
    ).pack(anchor="w", pady=(4, 10))
    var_resumo_selecao = tk.StringVar(value="")
    ttk.Label(direita, textvariable=var_resumo_selecao, style="Status.TLabel").pack(anchor="w", pady=(0, 8))
    preview = tk.Text(direita, height=18, wrap="word", state="disabled")
    preview.pack(fill="both", expand=True)

    def set_preview_text(texto: str) -> None:
        preview.configure(state="normal")
        preview.delete("1.0", tk.END)
        preview.insert("1.0", texto)
        preview.configure(state="disabled")

    def texto_detalhado_placa(numero: int) -> str:
        lote = placas[numero - 1]["lote"]  # numero eh 1-based
        linhas = [f"Placa {numero:03d}"]
        linhas.extend(resumo_lote(lote))
        return "\n".join(linhas)

    def aplicar_estado_marcacao_visivel() -> None:
        selecionadas_visiveis = [i for i, n in enumerate(exibidas) if n in selecionadas]
        listbox.selection_clear(0, tk.END)
        for idx in selecionadas_visiveis:
            listbox.selection_set(idx)

    def atualizar_contagem() -> None:
        var_contagem.set(
            f"Exibidas: {len(exibidas)} | Selecionadas: {len(selecionadas)} de {len(placas)}"
        )
        var_resumo_selecao.set(f"Faixas selecionadas: {resumir_intervalos_numericos(selecionadas)}")

    def atualizar_indicadores() -> None:
        com_conteudo = sum(1 for p in placas if int(p["qtd_validos"]) > 0)
        var_total_placas.set(str(len(placas)))
        var_selecionadas.set(str(len(selecionadas)))
        var_com_conteudo.set(str(com_conteudo))
        var_vazias.set(str(max(0, len(placas) - com_conteudo)))

    def reconstruir_lista(*_args) -> None:
        termo = normalizar_texto(var_filtro.get()).lower()
        status_atual = var_status_filtro.get()
        listbox.delete(0, tk.END)
        exibidas.clear()
        for p in placas:
            numero = int(p["numero"])
            lote = p["lote"]
            resumo = str(p["resumo"])
            tem_conteudo = int(p["qtd_validos"]) > 0
            texto_busca = " ".join(
                [
                    resumo.lower(),
                    " ".join(normalizar_texto(i.get("descricao", "")).lower() for i in lote),
                    " ".join(normalizar_preco_str(str(i.get("preco", ""))).lower() for i in lote),
                ]
            )
            if termo and termo not in texto_busca:
                continue
            if status_atual == "Com conteudo" and not tem_conteudo:
                continue
            if status_atual == "Vazias" and tem_conteudo:
                continue
            if status_atual == "Selecionadas" and numero not in selecionadas:
                continue
            if status_atual == "Nao selecionadas" and numero in selecionadas:
                continue
            item = f"Placa {numero:03d} | Itens validos: {int(p['qtd_validos'])} | {resumo}"
            listbox.insert(tk.END, item)
            exibidas.append(numero)

        aplicar_estado_marcacao_visivel()
        atualizar_contagem()
        atualizar_indicadores()
        atualizar_preview()

    def sincronizar_selecao_visivel(_event=None) -> None:
        visiveis = set(exibidas)
        selecionadas.difference_update(visiveis)
        indices = set(int(i) for i in listbox.curselection())
        for i, numero in enumerate(exibidas):
            if i in indices:
                selecionadas.add(numero)
        atualizar_contagem()
        atualizar_indicadores()
        atualizar_preview()

    def atualizar_preview(_event=None) -> None:
        cur = [int(i) for i in listbox.curselection()]
        if len(cur) == 1 and cur[0] < len(exibidas):
            numero = exibidas[cur[0]]
            set_preview_text(texto_detalhado_placa(numero))
            return
        if len(cur) > 1:
            nums = [exibidas[i] for i in cur[:6] if i < len(exibidas)]
            linhas = [f"{len(cur)} placas selecionadas (visiveis)."]
            for n in nums:
                linhas.append(f"- {texto_detalhado_placa(n).splitlines()[0]}")
            if len(cur) > 6:
                linhas.append("...")
            set_preview_text("\n".join(linhas))
            return
        if selecionadas:
            n = min(selecionadas)
            set_preview_text(texto_detalhado_placa(n))
            return
        set_preview_text("Nenhuma placa selecionada.")

    def marcar_todas() -> None:
        selecionadas.clear()
        for p in placas:
            selecionadas.add(int(p["numero"]))
        reconstruir_lista()

    def desmarcar_todas() -> None:
        selecionadas.clear()
        reconstruir_lista()

    def marcar_visiveis() -> None:
        for n in exibidas:
            selecionadas.add(int(n))
        reconstruir_lista()

    def desmarcar_visiveis() -> None:
        for n in exibidas:
            selecionadas.discard(int(n))
        reconstruir_lista()

    def marcar_apenas_nao_vazias() -> None:
        selecionadas.clear()
        for p in placas:
            if int(p["qtd_validos"]) > 0:
                selecionadas.add(int(p["numero"]))
        reconstruir_lista()

    def marcar_apenas_vazias() -> None:
        selecionadas.clear()
        for p in placas:
            if int(p["qtd_validos"]) <= 0:
                selecionadas.add(int(p["numero"]))
        reconstruir_lista()

    def inverter_selecao() -> None:
        numeros = {int(p["numero"]) for p in placas}
        atuais = set(selecionadas)
        selecionadas.clear()
        selecionadas.update(n for n in numeros if n not in atuais)
        reconstruir_lista()

    def aplicar_faixa(marcar: bool) -> None:
        try:
            numeros = parsear_intervalos_placas(var_intervalo.get(), len(placas))
        except ValueError as exc:
            messagebox.showerror("Faixa invalida", str(exc))
            return
        if not numeros:
            messagebox.showerror("Faixa obrigatoria", "Informe placas como 1-4, 7, 10-12.")
            return
        for numero in numeros:
            if marcar:
                selecionadas.add(numero)
            else:
                selecionadas.discard(numero)
        reconstruir_lista()

    def limpar_filtros() -> None:
        var_filtro.set("")
        var_status_filtro.set("Todas")
        var_intervalo.set("")
        reconstruir_lista()

    ttk.Button(acoes, text="Marcar Todas", command=marcar_todas).pack(side="left")
    ttk.Button(acoes, text="Desmarcar Todas", command=desmarcar_todas).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Marcar Exibidas", command=marcar_visiveis).pack(side="left", padx=(16, 0))
    ttk.Button(acoes, text="Desmarcar Exibidas", command=desmarcar_visiveis).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Somente Nao Vazias", command=marcar_apenas_nao_vazias).pack(side="left", padx=(16, 0))
    ttk.Button(acoes, text="Somente Vazias", command=marcar_apenas_vazias).pack(side="left", padx=(8, 0))
    ttk.Button(acoes, text="Inverter Selecao", command=inverter_selecao).pack(side="left", padx=(16, 0))

    ttk.Label(acoes_intervalo, text="Faixa:").pack(side="left")
    var_intervalo = tk.StringVar(value="")
    entrada_intervalo = ttk.Entry(acoes_intervalo, textvariable=var_intervalo, width=18)
    entrada_intervalo.pack(side="left", padx=(8, 8))
    ttk.Label(acoes_intervalo, text="Ex.: 1-4, 7, 10-12", style="Status.TLabel").pack(side="left")
    ttk.Button(acoes_intervalo, text="Marcar Faixa", command=lambda: aplicar_faixa(True)).pack(
        side="left", padx=(16, 0)
    )
    ttk.Button(acoes_intervalo, text="Desmarcar Faixa", command=lambda: aplicar_faixa(False)).pack(
        side="left", padx=(8, 0)
    )
    ttk.Button(acoes_intervalo, text="Limpar Filtros", command=limpar_filtros).pack(side="left", padx=(16, 0))

    rodape = ttk.Frame(root, style="Page.TFrame", padding=(16, 12))
    rodape.pack(fill="x")
    var_desligar_ao_final = tk.BooleanVar(value=bool(desligar_ao_final_inicial))
    ttk.Checkbutton(
        rodape,
        text="Desligar o PC ao final da impressao",
        variable=var_desligar_ao_final,
    ).pack(side="left")

    def confirmar() -> None:
        if not selecionadas:
            messagebox.showerror("Selecao obrigatoria", "Selecione ao menos 1 placa para produzir.")
            return
        resultado["selecionadas"] = sorted(selecionadas)
        resultado["desligar_ao_final"] = bool(var_desligar_ao_final.get())
        resultado["ok"] = True
        root.destroy()

    def cancelar() -> None:
        if messagebox.askyesno("Cancelar", "Cancelar a producao?"):
            resultado["ok"] = False
            root.destroy()

    ttk.Button(rodape, text="Cancelar", command=cancelar).pack(side="right")
    ttk.Button(rodape, text="Iniciar Producao", command=confirmar).pack(side="right", padx=(0, 8))

    listbox.bind("<<ListboxSelect>>", sincronizar_selecao_visivel)
    var_filtro.trace_add("write", reconstruir_lista)
    var_status_filtro.trace_add("write", reconstruir_lista)
    root.bind("<Control-Return>", lambda _event: confirmar())
    root.bind("<Escape>", lambda _event: cancelar())
    root.protocol("WM_DELETE_WINDOW", cancelar)

    reconstruir_lista()
    entrada_filtro.focus_set()
    root.mainloop()

    if not bool(resultado.get("ok")):
        return None
    return (
        list(resultado.get("selecionadas", [])),  # type: ignore[arg-type]
        bool(resultado.get("desligar_ao_final", False)),
    )


def confirmar_impressao_placa_tela(
    numero_placa: int,
    ordem_execucao: int,
    total_lotes: int,
    lote: List[Dict[str, str]],
) -> str:
    """
    Retorna:
      - "proximo": seguir para a proxima placa
      - "reimprimir": repetir impressao desta placa
      - "cancelar": encerrar producao
    """
    mensagem = (
        f"Placa {numero_placa} ({ordem_execucao}/{total_lotes})\n\n"
        "A impressao saiu correta?\n"
        "Sim = proxima placa\n"
        "Nao = reimprimir esta placa\n"
        "Cancelar = encerrar producao\n\n"
        + "\n".join(resumo_lote(lote))
    )

    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        try:
            root.attributes("-topmost", True)
        except Exception:
            pass
        resposta = messagebox.askyesnocancel("Confirmar Impressao", mensagem, parent=root)
        root.destroy()
        if resposta is True:
            return "proximo"
        if resposta is False:
            return "reimprimir"
        return "cancelar"
    except Exception:
        print("\n" + mensagem)
        while True:
            resp = input("Confirmar [S=proxima / R=reimprimir / C=cancelar]: ").strip().lower()
            if resp in ("s", "sim", "y", "yes", ""):
                return "proximo"
            if resp in ("r", "reimprimir", "n", "nao", "não"):
                return "reimprimir"
            if resp in ("c", "cancelar", "q", "quit"):
                return "cancelar"
            print("Opcao invalida. Digite S, R ou C.")


def executar_lotes(
    arquivo_cdr: Path,
    salvar_em: Optional[Path],
    lotes_info: List[Tuple[int, List[Dict[str, str]]]],
    imprimir: bool,
    copias: int,
    impressora: Optional[str],
    pausa_segundos: float,
    salvar_documento: bool,
    confirmar_impressao: bool,
    modo_rapido_inteligente: bool,
    progress_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Dict[str, object]:
    config_base = {
        "fonte_descricao_unidade": "TangoSans",
        "tamanho_descricao_min": 24,
        "tamanho_descricao_max": 38,
        "largura_max_descricao_cm": 5.0,
        "altura_max_descricao_cm": 1.9,
        "max_linhas_descricao": 2,
        "preferir_descricao_duas_linhas": True,
        "descricao_palavras_min_duas_linhas": 3,
        "tamanho_descricao_min_emergencia": 16,
        "tamanho_preco_min": 90,
        "tamanho_preco_max": 170,
        "fator_tamanho_centavos": 0.37,
        "deslocamento_vertical_centavos": 118,
        "fator_largura_preco_auto": 0.78,
        "fator_largura_preco_int_auto": 0.60,
        "alinhar_centavos_preco": True,
        "gap_preco_centavos_cm": 0.10,
        "centralizar_preco": True,
        "altura_max_preco_cm": 3.7,
        "altura_max_preco_int_cm": 3.7,
        "reposicionar_campos": False,
        "fator_tamanho_unidade": 0.42,
        "tamanho_unidade_max": 28,
        "alinhar_unidade_centavos": True,
        "ajustar_colisao_unidade": False,
        "deslocamento_x_unidade_cm": 0.02,
        "gap_unidade_min_cm": 0.252,
        "gap_unidade_fator_altura": 0.0,
        "permitir_produto_vazio": True,
    }

    arquivo_trabalho = arquivo_cdr
    if salvar_documento and salvar_em is not None:
        arquivo_trabalho = salvar_em

    total_lotes = len(lotes_info)
    usar_sessao_continua = bool(modo_rapido_inteligente and total_lotes > 1)
    tempos_tentativas: List[float] = []
    inicio_total = time.perf_counter()
    abriu_documento_pela_automacao = False

    if callable(progress_callback):
        progress_callback(
            {
                "event": "started",
                "total_plates": total_lotes,
                "completed_plates": 0,
                "current_plate": 0,
                "current_plate_label": "",
                "current_product": "",
            }
        )

    try:
        for ordem_execucao, (numero_placa, lote) in enumerate(lotes_info, start=1):
            config_lote = dict(config_base)
            config_lote["produtos"] = lote
            produto_atual = resumo_lote_curto(lote, max_itens=1)

            if salvar_documento:
                if ordem_execucao == 1:
                    cdr_origem = arquivo_cdr
                    salvar_destino = salvar_em
                else:
                    cdr_origem = arquivo_trabalho
                    salvar_destino = None
            else:
                cdr_origem = arquivo_cdr
                salvar_destino = None

            print(f"\nPlaca {numero_placa} ({ordem_execucao}/{total_lotes})")
            for linha in resumo_lote(lote):
                print(linha)

            if callable(progress_callback):
                elapsed_seconds = max(0.0, time.perf_counter() - inicio_total)
                eta_seconds = None
                if ordem_execucao > 1:
                    media_placa = elapsed_seconds / max(1, ordem_execucao - 1)
                    eta_seconds = max(0.0, media_placa * (total_lotes - (ordem_execucao - 1)))
                progress_callback(
                    {
                        "event": "plate_started",
                        "total_plates": total_lotes,
                        "completed_plates": ordem_execucao - 1,
                        "execution_order": ordem_execucao,
                        "current_plate": numero_placa,
                        "current_plate_label": f"Placa {numero_placa} ({ordem_execucao}/{total_lotes})",
                        "current_product": produto_atual,
                        "lote": [dict(item) for item in lote],
                        "elapsed_seconds": elapsed_seconds,
                        "eta_seconds": eta_seconds,
                    }
                )

            while True:
                inicio_tentativa = time.perf_counter()
                abriu_documento = corel.atualizar_documento(
                    config=config_lote,
                    cdr_path=cdr_origem,
                    salvar_em=salvar_destino,
                    imprimir=imprimir,
                    copias=max(1, copias),
                    impressora=impressora,
                    salvar_documento=salvar_documento,
                    fechar_documento=(not usar_sessao_continua),
                )
                tempos_tentativas.append(max(0.0, time.perf_counter() - inicio_tentativa))
                if usar_sessao_continua and abriu_documento:
                    abriu_documento_pela_automacao = True

                if not imprimir or not confirmar_impressao:
                    break

                acao = confirmar_impressao_placa_tela(
                    numero_placa=numero_placa,
                    ordem_execucao=ordem_execucao,
                    total_lotes=total_lotes,
                    lote=lote,
                )
                if acao == "proximo":
                    break
                if acao == "reimprimir":
                    continue
                raise ProducaoCancelada("Producao cancelada pelo usuario apos etapa de impressao.")

            if callable(progress_callback):
                elapsed_seconds = max(0.0, time.perf_counter() - inicio_total)
                media_placa = elapsed_seconds / max(1, ordem_execucao)
                eta_seconds = max(0.0, media_placa * (total_lotes - ordem_execucao))
                progress_callback(
                    {
                        "event": "plate_completed",
                        "total_plates": total_lotes,
                        "completed_plates": ordem_execucao,
                        "execution_order": ordem_execucao,
                        "current_plate": numero_placa,
                        "current_plate_label": f"Placa {numero_placa} ({ordem_execucao}/{total_lotes})",
                        "current_product": produto_atual,
                        "lote": [dict(item) for item in lote],
                        "elapsed_seconds": elapsed_seconds,
                        "eta_seconds": eta_seconds,
                    }
                )

            if pausa_segundos > 0 and ordem_execucao < total_lotes:
                time.sleep(pausa_segundos)
    finally:
        if usar_sessao_continua and abriu_documento_pela_automacao:
            try:
                corel.fechar_documento_por_caminho(
                    cdr_path=arquivo_trabalho,
                    salvar_documento=salvar_documento,
                )
            except Exception as exc:
                print(f"Aviso: nao foi possivel fechar automaticamente o documento no Corel ({exc}).")

    total_segundos = max(0.0, time.perf_counter() - inicio_total)
    media_tentativa = (sum(tempos_tentativas) / len(tempos_tentativas)) if tempos_tentativas else 0.0
    return {
        "total_segundos": total_segundos,
        "media_segundos_tentativa": media_tentativa,
        "tempos_tentativas": tempos_tentativas,
        "modo_rapido_ativo": usar_sessao_continua,
        "total_lotes": total_lotes,
    }


def main() -> int:
    carregar_variaveis_dotenv()
    args = parser_args().parse_args()

    arquivo_cdr = Path(args.arquivo_cdr).resolve()
    salvar_em = Path(args.salvar_em).resolve() if args.salvar_em else None
    aprendizado_path = Path(args.arquivo_aprendizado).resolve()
    velocidade_path = Path(args.arquivo_velocidade).resolve()
    corretor_acentos_path = Path(args.arquivo_corretor_acentos).resolve()
    origem_entrada = ""

    if not arquivo_cdr.exists():
        print(f"Arquivo CDR nao encontrado: {arquivo_cdr}", file=sys.stderr)
        return 2

    github_repo_acesso = normalizar_texto(os.environ.get(GITHUB_REPO_ENV, DEFAULT_GITHUB_LOG_REPO)) or DEFAULT_GITHUB_LOG_REPO
    github_branch_acesso = normalizar_texto(os.environ.get(GITHUB_BRANCH_ENV, DEFAULT_GITHUB_LOG_BRANCH)) or DEFAULT_GITHUB_LOG_BRANCH
    github_token_acesso = normalizar_texto(os.environ.get(GITHUB_TOKEN_ENV, ""))
    github_aprendizado_path = (
        normalizar_texto(os.environ.get(GITHUB_APRENDIZADO_PATH_ENV, DEFAULT_GITHUB_APRENDIZADO_PATH))
        or DEFAULT_GITHUB_APRENDIZADO_PATH
    )
    usuarios_acesso, origem_usuarios = carregar_usuarios_acesso_github(
        token=github_token_acesso or None,
        repo=DEFAULT_GITHUB_LOG_REPO,
        caminho_arquivo=DEFAULT_GITHUB_USERS_PATH,
        branch=DEFAULT_GITHUB_LOG_BRANCH,
    )
    if not usuarios_acesso:
        print(
            "Acesso bloqueado: nenhum usuario valido foi carregado de "
            f"{origem_usuarios}. Atualize https://github.com/PopularAtacarejo/Placas/blob/main/usuarios.json "
            "com campo senha_hash (PBKDF2)."
        )
        return 2
    usuario_logado = carregar_sessao_login_24h_valida(usuarios_acesso)
    if usuario_logado is None:
        usuario_logado = autenticar_acesso_navegador(
            usuarios=usuarios_acesso,
            source_label=origem_usuarios,
        )
    if usuario_logado is None:
        print("[LOGIN] Acesso cancelado pelo usuario.")
        return 0
    if not usuario_pode_gerar_placas(usuario_logado):
        print("[LOGIN] Acesso bloqueado: o nivel deste usuario nao permite geracao de placas.")
        return 2
    nome_usuario_logado = normalizar_texto(str(usuario_logado.get("nome") or usuario_logado.get("usuario") or "usuario"))
    perfil_usuario_logado = normalizar_texto(str(usuario_logado.get("perfil") or ""))
    mensagem_login = "Login liberado para " + nome_usuario_logado
    if perfil_usuario_logado:
        mensagem_login += f" ({perfil_usuario_logado})"
    if bool(usuario_logado.get("sessao_24h")):
        expira_em = normalizar_texto(str(usuario_logado.get("sessao_expira_em") or ""))
        if expira_em:
            mensagem_login += f" via sessao de 24 horas ate {expira_em}."
        else:
            mensagem_login += " via sessao de 24 horas."
    print(f"[LOGIN] {mensagem_login}")

    corretor_acentos: Dict[str, str] = {}
    if not args.sem_corretor_acentos:
        corretor_acentos = carregar_corretor_acentos(corretor_acentos_path)

    itens_aprendidos: Dict[str, Dict[str, object]] = {}
    if not args.sem_aprendizado:
        try:
            itens_aprendidos = carregar_aprendizado_github(
                token=github_token_acesso or None,
                repo=github_repo_acesso,
                caminho_arquivo=github_aprendizado_path,
                branch=github_branch_acesso,
            )
            print(f"[IA] Aprendizado: base carregada de {github_repo_acesso}/{github_aprendizado_path}.")
        except Exception as exc:
            itens_aprendidos = carregar_aprendizado(aprendizado_path)
            print(f"[AVISO] Nao foi possivel carregar o aprendizado remoto ({exc}). Usando base local {aprendizado_path}.")

    produtos: List[Dict[str, str]] = []
    produtos_originais_extraidos: List[Dict[str, str]] = []
    stats_processamento_inicial: Dict[str, int] = {}
    entrada_precarregada = False

    try:
        if args.texto_bruto is not None and normalizar_texto(args.texto_bruto):
            origem_entrada = "texto bruto (argumento)"
            produtos_originais_extraidos = extrair_texto_bruto_lista(args.texto_bruto)
            produtos, stats_processamento_inicial = processar_produtos_inteligentes(
                produtos=produtos_originais_extraidos,
                corretor_acentos=corretor_acentos,
                itens_aprendidos=itens_aprendidos,
                usar_aprendizado=(not args.sem_aprendizado),
                usar_ia_local=bool(args.usar_ia_local),
                limite_itens_ia_local=max(0, int(args.max_itens_ia_local)),
                modelo_ia_local=str(args.modelo_ia_local or DEFAULT_OLLAMA_MODEL),
                timeout_ia_local=max(1.0, float(args.timeout_ia_local)),
            )
            entrada_precarregada = True
        elif args.entrada:
            entrada = Path(args.entrada).resolve()
            if not entrada.exists():
                print(f"Entrada nao encontrada: {entrada}", file=sys.stderr)
                return 2
            origem_entrada = str(entrada)
            produtos_originais_extraidos = extrair_produtos(entrada)
            produtos, stats_processamento_inicial = processar_produtos_inteligentes(
                produtos=produtos_originais_extraidos,
                corretor_acentos=corretor_acentos,
                itens_aprendidos=itens_aprendidos,
                usar_aprendizado=(not args.sem_aprendizado),
                usar_ia_local=bool(args.usar_ia_local),
                limite_itens_ia_local=max(0, int(args.max_itens_ia_local)),
                modelo_ia_local=str(args.modelo_ia_local or DEFAULT_OLLAMA_MODEL),
                timeout_ia_local=max(1.0, float(args.timeout_ia_local)),
            )
            entrada_precarregada = True
        elif args.sem_tela_unidade:
            selecao = selecionar_entrada_tela()
            if selecao is None:
                print("Processo cancelado pelo usuario.")
                return 0
            tipo_entrada = selecao.get("tipo", "")
            valor_entrada = str(selecao.get("valor", ""))
            if tipo_entrada == "texto":
                origem_entrada = "texto bruto (colado)"
                produtos_originais_extraidos = extrair_texto_bruto_lista(valor_entrada)
            else:
                entrada = Path(valor_entrada).resolve()
                if not entrada.exists():
                    print(f"Entrada nao encontrada: {entrada}", file=sys.stderr)
                    return 2
                origem_entrada = str(entrada)
                produtos_originais_extraidos = extrair_produtos(entrada)
            produtos, stats_processamento_inicial = processar_produtos_inteligentes(
                produtos=produtos_originais_extraidos,
                corretor_acentos=corretor_acentos,
                itens_aprendidos=itens_aprendidos,
                usar_aprendizado=(not args.sem_aprendizado),
                usar_ia_local=bool(args.usar_ia_local),
                limite_itens_ia_local=max(0, int(args.max_itens_ia_local)),
                modelo_ia_local=str(args.modelo_ia_local or DEFAULT_OLLAMA_MODEL),
                timeout_ia_local=max(1.0, float(args.timeout_ia_local)),
            )
            entrada_precarregada = True
    except EntradaInvalida as exc:
        print(f"[ERRO] Falha na leitura da entrada: {exc}", file=sys.stderr)
        return 1

    if entrada_precarregada:
        if stats_processamento_inicial.get("acentos", 0) > 0:
            print(
                f"[TEXTO] Corretor de acentos: {stats_processamento_inicial.get('acentos', 0)} descricoes ajustadas com {corretor_acentos_path}."
            )
        if stats_processamento_inicial.get("aprendizado", 0) > 0:
            print(
                f"[IA] Aprendizado: {stats_processamento_inicial.get('aprendizado', 0)} sugestoes aplicadas automaticamente."
            )
        if stats_processamento_inicial.get("ia_local", 0) > 0:
            print(
                f"[IA] Ollama: {stats_processamento_inicial.get('ia_local', 0)} descricao(oes) limpas automaticamente."
            )
        if stats_processamento_inicial.get("acentos_pos_ia", 0) > 0:
            print(
                f"[TEXTO] Corretor de acentos: {stats_processamento_inicial.get('acentos_pos_ia', 0)} ajustes apos a IA local."
            )
        if stats_processamento_inicial.get("aprendizado_pos_ia", 0) > 0:
            print(
                "[IA] Aprendizado: "
                f"{stats_processamento_inicial.get('aprendizado_pos_ia', 0)} sugestoes reaplicadas apos a limpeza local."
            )

    placas_selecionadas_da_revisao: Optional[List[int]] = None
    desligar_ao_final = bool(args.desligar_ao_final)
    estado_revisao: Dict[str, Any] = {}
    if not args.sem_tela_unidade:
        def _analisar_entrada_callback(payload: Dict[str, Any]) -> Dict[str, Any]:
            return analisar_entrada_web_payload(
                payload=payload,
                corretor_acentos=corretor_acentos,
                itens_aprendidos=itens_aprendidos,
                usar_aprendizado=(not args.sem_aprendizado),
                usar_ia_local=bool(args.usar_ia_local),
                limite_itens_ia_local=max(0, int(args.max_itens_ia_local)),
                modelo_ia_local=str(args.modelo_ia_local or DEFAULT_OLLAMA_MODEL),
                timeout_ia_local=max(1.0, float(args.timeout_ia_local)),
            )

        revisao = revisar_produtos_interface(
            produtos,
            permitir_selecao_placas=(not args.sem_tela_placas),
            desligar_ao_final_inicial=desligar_ao_final,
            permitir_analise_entrada=True,
            analisar_entrada_callback=_analisar_entrada_callback,
            estado_revisao=estado_revisao,
            produtos_originais_iniciais=produtos_originais_extraidos,
            usuario_logado=usuario_logado,
            usuarios_acesso=usuarios_acesso,
        )
        if revisao is None:
            encerrar_sessao_web_revisao(estado_revisao)
            print("Processo cancelado pelo usuario.")
            return 0
        produtos, placas_selecionadas_da_revisao, desligar_ao_final = revisao
        if isinstance(estado_revisao.get("produtos_originais_extraidos"), list):
            produtos_originais_extraidos = [
                dict(item)
                for item in estado_revisao.get("produtos_originais_extraidos", [])
                if isinstance(item, dict)
            ]
        origem_revisao = normalizar_texto(str(estado_revisao.get("origem_entrada", "")))
        if origem_revisao:
            origem_entrada = origem_revisao
        usuario_logado_revisao = estado_revisao.get("usuario_logado")
        if isinstance(usuario_logado_revisao, dict) and usuario_logado_revisao:
            usuario_logado = dict(usuario_logado_revisao)
        if corretor_acentos:
            produtos, qtd_acentos_pos_tela = aplicar_corretor_acentos_produtos(produtos, corretor_acentos)
            if qtd_acentos_pos_tela > 0:
                print(f"[TEXTO] Corretor de acentos: {qtd_acentos_pos_tela} descricoes ajustadas apos revisao.")

    if not args.sem_aprendizado:
        qtd_novos = registrar_aprendizado(
            originais=produtos_originais_extraidos,
            revisados=produtos,
            itens_aprendidos=itens_aprendidos,
        )
        if qtd_novos > 0:
            aprendizado_salvo = False
            if github_token_acesso:
                try:
                    salvar_aprendizado_github(
                        token=github_token_acesso,
                        repo=github_repo_acesso,
                        caminho_arquivo=github_aprendizado_path,
                        branch=github_branch_acesso,
                        itens=itens_aprendidos,
                    )
                    aprendizado_salvo = True
                    print(
                        "[IA] Aprendizado: "
                        f"{qtd_novos} correcoes memorizadas em {github_repo_acesso}/{github_aprendizado_path}."
                    )
                except Exception as exc:
                    print(f"[AVISO] Nao foi possivel gravar o aprendizado remoto ({exc}).")
            if not aprendizado_salvo:
                aprendizado_path.parent.mkdir(parents=True, exist_ok=True)
                salvar_aprendizado(aprendizado_path, itens_aprendidos)
                print(f"[IA] Aprendizado: {qtd_novos} correcoes memorizadas em {aprendizado_path}.")

    lotes = montar_lotes(produtos, tamanho_lote=4)
    if not lotes:
        print("[ERRO] Nenhum produto valido foi encontrado.", file=sys.stderr)
        return 1

    lotes_info: List[Tuple[int, List[Dict[str, str]]]] = [
        (i, lote) for i, lote in enumerate(lotes, start=1)
    ]

    if placas_selecionadas_da_revisao is not None:
        set_sel = set(int(i) for i in placas_selecionadas_da_revisao)
        lotes_info = [(i, lote) for i, lote in lotes_info if i in set_sel]
        if not lotes_info:
            print("Nenhuma placa selecionada.", file=sys.stderr)
            return 1
    elif not args.sem_tela_placas:
        selecao_placas = selecionar_placas_tela(lotes, desligar_ao_final_inicial=desligar_ao_final)
        if selecao_placas is None:
            print("Processo cancelado pelo usuario.")
            return 0
        placas_selecionadas, desligar_ao_final = selecao_placas
        set_sel = set(int(i) for i in placas_selecionadas)
        lotes_info = [(i, lote) for i, lote in lotes_info if i in set_sel]
        if not lotes_info:
            print("Nenhuma placa selecionada.", file=sys.stderr)
            return 1

    salvar_documento = bool(args.salvar_cdr)
    if salvar_documento:
        if not args.sem_copia and salvar_em is None:
            salvar_em = construir_saida_default(arquivo_cdr)
    else:
        salvar_em = None

    imprimir = not bool(args.nao_imprimir)
    confirmar_impressao = bool(imprimir and (not args.sem_confirmacao_impressao))
    template_key = chave_template_execucao(arquivo_cdr)
    perfis_velocidade: Dict[str, Dict[str, object]] = {}
    perfil_velocidade: Dict[str, object] = {}
    if not args.sem_aprendizado:
        perfis_velocidade = carregar_perfil_velocidade(velocidade_path)
        perfil_velocidade = perfis_velocidade.get(template_key, {})

    modo_rapido_inteligente = (
        (not args.sem_modo_rapido_inteligente) and decidir_modo_rapido_inteligente(perfil_velocidade)
    )
    github_log_token = normalizar_texto(os.environ.get(GITHUB_TOKEN_ENV, ""))
    github_log_repo = DEFAULT_GITHUB_LOG_REPO
    github_log_path = DEFAULT_GITHUB_LOG_PATH
    github_log_branch = DEFAULT_GITHUB_LOG_BRANCH
    github_log_habilitado = bool(github_log_token)
    github_log_estado: Dict[str, Any] = {
        "habilitado": github_log_habilitado,
        "token": github_log_token,
        "repo": github_log_repo,
        "path": github_log_path,
        "branch": github_log_branch,
        "falha_emitida": False,
    }

    print(f"Entrada: {origem_entrada}")
    print(f"Produtos extraidos: {len(produtos)}")
    print(f"Placas selecionadas para producao: {len(lotes_info)}/{len(lotes)}")
    print(
        "Persistencia CDR: "
        + ("habilitada (salva alteracoes no arquivo)" if salvar_documento else "desabilitada (somente impressao)")
    )
    print(
        "Corretor de acentos: "
        + ("habilitado" if (not args.sem_corretor_acentos and bool(corretor_acentos)) else "desabilitado")
    )
    print(
        "IA local: "
        + (
            f"habilitada ({args.modelo_ia_local}, max {max(0, int(args.max_itens_ia_local))} item(ns))"
            if args.usar_ia_local
            else "desabilitada"
        )
    )
    print(
        "Modo rapido inteligente: "
        + ("habilitado" if modo_rapido_inteligente else "desabilitado")
    )
    if args.sem_log_github:
        print("Aviso: --sem-log-github ignorado. O log remoto das placas e obrigatorio.")
    if github_log_habilitado:
        print(f"Log GitHub: habilitado em {github_log_repo}/{github_log_path} ({github_log_branch}).")
    else:
        print(f"Log GitHub: bloqueado (defina {GITHUB_TOKEN_ENV} para gravar as placas no historico remoto).")
        return 2
    if isinstance(perfil_velocidade, dict):
        media_hist = perfil_velocidade.get("media_segundos_por_placa")
        try:
            media_hist_f = float(media_hist)
        except Exception:
            media_hist_f = 0.0
        if media_hist_f > 0:
            print(f"Historico de velocidade: media {media_hist_f:.2f}s por placa.")
    if imprimir:
        print(f"Impressao: habilitada | Copias por lote: {max(1, args.copias)}")
        print(
            "Confirmacao por placa: "
            + ("habilitada" if confirmar_impressao else "desabilitada (--sem-confirmacao-impressao)")
        )
        print(
            "Desligar ao final: "
            + ("habilitado" if desligar_ao_final else "desabilitado")
        )
        if args.impressora:
            print(f"Impressora: {args.impressora}")
    else:
        print("Impressao: desabilitada (--nao-imprimir)")
        if desligar_ao_final:
            print("Aviso: desligamento automatico ignorado porque a impressao esta desabilitada.")
            desligar_ao_final = False

    estatisticas_execucao: Dict[str, object] = {}
    def _progress_callback(payload: Dict[str, Any]) -> None:
        progress_state = estado_revisao.get("web_progress_state")
        now_ts = time.time()
        event = str(payload.get("event", "") or "")
        if event == "started" and isinstance(progress_state, dict):
            progress_state.update(
                {
                    "phase": "running",
                    "message": "Producao iniciada.",
                    "started_at": now_ts,
                    "finished_at": None,
                    "updated_at": now_ts,
                    "total_plates": int(payload.get("total_plates", 0) or 0),
                    "completed_plates": 0,
                    "current_plate": 0,
                    "current_plate_label": "",
                    "current_product": "",
                    "eta_seconds": None,
                    "elapsed_seconds": 0.0,
                }
            )
            return
        if event in {"plate_started", "plate_completed"} and isinstance(progress_state, dict):
            progress_state["phase"] = "running"
            progress_state["message"] = (
                "Produzindo placa atual..." if event == "plate_started" else "Placa concluida. Seguindo fluxo..."
            )
            progress_state["updated_at"] = now_ts
            progress_state["total_plates"] = int(payload.get("total_plates", progress_state.get("total_plates", 0)) or 0)
            progress_state["completed_plates"] = int(payload.get("completed_plates", progress_state.get("completed_plates", 0)) or 0)
            progress_state["current_plate"] = int(payload.get("current_plate", progress_state.get("current_plate", 0)) or 0)
            progress_state["current_plate_label"] = str(payload.get("current_plate_label", progress_state.get("current_plate_label", "")) or "")
            progress_state["current_product"] = str(payload.get("current_product", progress_state.get("current_product", "")) or "")
            progress_state["eta_seconds"] = payload.get("eta_seconds")
            progress_state["elapsed_seconds"] = float(payload.get("elapsed_seconds", progress_state.get("elapsed_seconds", 0.0)) or 0.0)
        if event == "plate_completed" and github_log_estado.get("habilitado"):
            try:
                lote_payload = payload.get("lote")
                lote_log = [dict(item) for item in lote_payload] if isinstance(lote_payload, list) else []
                registro = montar_registro_placa_github(
                    numero_placa=int(payload.get("current_plate", 0) or 0),
                    lote=lote_log,
                    usuario_logado=usuario_logado,
                )
                registrar_placa_concluida_github(
                    token=str(github_log_estado.get("token", "")),
                    repo=str(github_log_estado.get("repo", DEFAULT_GITHUB_LOG_REPO)),
                    caminho_arquivo=str(github_log_estado.get("path", DEFAULT_GITHUB_LOG_PATH)),
                    branch=str(github_log_estado.get("branch", DEFAULT_GITHUB_LOG_BRANCH)),
                    registro=registro,
                )
            except Exception as exc:
                if not bool(github_log_estado.get("falha_emitida", False)):
                    github_log_estado["falha_emitida"] = True
                raise RuntimeError(f"Nao foi possivel registrar a placa no GitHub ({exc}).")

    try:
        estatisticas_execucao = executar_lotes(
            arquivo_cdr=arquivo_cdr,
            salvar_em=salvar_em,
            lotes_info=lotes_info,
            imprimir=imprimir,
            copias=max(1, args.copias),
            impressora=args.impressora,
            pausa_segundos=max(0.0, float(args.pausa_segundos)),
            salvar_documento=salvar_documento,
            confirmar_impressao=confirmar_impressao,
            modo_rapido_inteligente=modo_rapido_inteligente,
            progress_callback=_progress_callback,
        )
    except ProducaoCancelada as exc:
        progress_state = estado_revisao.get("web_progress_state")
        if isinstance(progress_state, dict):
            progress_state.update(
                {
                    "phase": "cancelled",
                    "message": str(exc),
                    "finished_at": time.time(),
                    "updated_at": time.time(),
                }
            )
            time.sleep(1.0)
            encerrar_sessao_web_revisao(estado_revisao)
        print(str(exc))
        return 0
    except Exception as exc:
        progress_state = estado_revisao.get("web_progress_state")
        if isinstance(progress_state, dict):
            progress_state.update(
                {
                    "phase": "error",
                    "message": str(exc),
                    "finished_at": time.time(),
                    "updated_at": time.time(),
                }
            )
            time.sleep(1.0)
            encerrar_sessao_web_revisao(estado_revisao)
        if not args.sem_aprendizado:
            registrar_resultado_velocidade(
                perfis=perfis_velocidade,
                template_key=template_key,
                modo_rapido_ativo=modo_rapido_inteligente,
                tempos_placa=[],
                sucesso=False,
                erro=str(exc),
            )
            velocidade_path.parent.mkdir(parents=True, exist_ok=True)
            salvar_perfil_velocidade(velocidade_path, perfis_velocidade)
        print(f"Falha ao atualizar/imprimir lotes: {exc}", file=sys.stderr)
        return 1

    if not args.sem_aprendizado:
        tempos_tentativas: List[float] = []
        for t in estatisticas_execucao.get("tempos_tentativas", []):
            try:
                tempos_tentativas.append(float(t))
            except Exception:
                continue
        perfil_atualizado = registrar_resultado_velocidade(
            perfis=perfis_velocidade,
            template_key=template_key,
            modo_rapido_ativo=bool(estatisticas_execucao.get("modo_rapido_ativo", False)),
            tempos_placa=tempos_tentativas,
            sucesso=True,
        )
        velocidade_path.parent.mkdir(parents=True, exist_ok=True)
        salvar_perfil_velocidade(velocidade_path, perfis_velocidade)

        try:
            media_hist = float(perfil_atualizado.get("media_segundos_por_placa", 0.0))
        except Exception:
            media_hist = 0.0
        if media_hist > 0:
            print(
                "Aprendizado de velocidade: "
                f"media historica atualizada para {media_hist:.2f}s por placa em {velocidade_path}."
            )

    print("\nProcessamento finalizado.")
    progress_state = estado_revisao.get("web_progress_state")
    if isinstance(progress_state, dict):
        progress_state.update(
            {
                "phase": "finished",
                "message": "Processamento finalizado.",
                "finished_at": time.time(),
                "updated_at": time.time(),
                "completed_plates": int(progress_state.get("total_plates", len(lotes_info)) or len(lotes_info)),
                "eta_seconds": 0.0,
            }
        )
        time.sleep(1.0)
    encerrar_sessao_web_revisao(estado_revisao)
    if desligar_ao_final and imprimir and lotes_info:
        ok_desligamento, mensagem_desligamento = agendar_desligamento_windows(segundos=60)
        print(mensagem_desligamento)
        if not ok_desligamento:
            return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
